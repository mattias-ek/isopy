import functools
import os as _os
import warnings

from isopy import core
import csv as csv
import datetime as dt
import numpy as np
import chardet
import openpyxl
import pyperclip
from openpyxl import load_workbook
import itertools
import io
import zipfile, os
import numpy as np


__all__ = ['read_exp',
           'read_csv', 'write_csv',
           'read_xlsx', 'write_xlsx',
           'read_clipboard', 'write_clipboard',
           'new_archive', 'load_archive']

import isopy.checks

NAN_STRINGS = 'nan #NA #N/A N/A NA =NA() =na()'.split()

def rows_to_data(data, has_keys, keys_in_first, description = None):
    if description:
        keys_in_first = description.get('keys_in_first', keys_in_first)

        output_type = description.get('type', None)
        if  output_type == 'array':
            has_keys = True
        elif output_type == 'ndarray':
            has_keys = False
    else:
        output_type = None
    
    if len(data) == 0 or len(data[0])==0:
        if has_keys is not True and keys_in_first is None:
            return [[]]
        else:
            return dict()

    # Everything needs to be converted to nan before it gets here
    if has_keys is not False and keys_in_first is None:
        r = [type(v) is str for v in data[0]]
        c = [type(d[0]) is str for d in data]

        if False not in c and False not in r:
            # Both are strings
            try:
                float(data[0][0])
            except ValueError:
                f = True
            else:
                f = False

            try:
                [float(v) for v in data[0][1:]]
            except ValueError:
                r = True
            else:
                r = False

            try:
                [float(v[0]) for v in data[1:]]
            except ValueError:
                c = True
            else:
                c = False

            if has_keys is None and not c and not r:
                if f and (len(data)) == 1:
                    keys_in_first = 'c'
                elif f and len(data[0]) == 1:
                    keys_in_first = 'r'
                elif f:
                    # Both the first row and the first
                    warnings.warn('Both the first row and the first column contain strings. Using first row as keys')
                    keys_in_first = 'r'
                    #raise ValueError('Unable to determine whether the first rows or columns contains the keys')
                else:
                    # All values in the first row and the first column can be converted to float
                    has_keys = False
            elif c and not r:
                keys_in_first = 'c'
            elif r and not c:
                keys_in_first = 'r'
            else:
                # Both the first row and the first column contains at least one value that cant be converted to float
                warnings.warn('Both the first row and the first column contain strings. Using first row as keys')
                keys_in_first = 'r'
                #raise ValueError('Unable to determine whether the first rows or columns contains the keys')

        elif False not in c:
            # Only c contains string
            keys_in_first = 'c'
        elif False not in r:
            # Only r contains strings
            keys_in_first = 'r'
        else:
            # Neither contain only strings
            has_keys=False
    
    if has_keys is False:
        output = data
    elif keys_in_first == 'c':
        output = {v[0]: v[1:] for v in data}
    elif keys_in_first == 'r':
        output = {v[0]: v[1:] for v in zip(*data)}
    else:
        raise ValueError(f'Unknown value for "keys_in_first" {keys_in_first}')
    
    if output_type is None:
        return output
    
    elif output_type == 'array':
        ndim = description.get('ndim', '1')
        if ndim =='0':
            output = {k: v[0] for k, v in output.items()}
            
        flavours = description.get('flavour', None)
        if type(flavours) is not list:
            flavours = [flavours] * len(output)
            
        datatypes = description.get('dtype', None)
        if type(datatypes) is not list:
            datatypes = [datatypes] * len(output)
        
        output = {isopy.keystring(k, flavour=flavours[i]): np.array(v, dtype=datatypes[i]) 
                for i, (k,v) in enumerate(output.items())}
        
        return isopy.array(output)
        
    elif output_type == 'refval':
        # Not possible
        flavours = description.get('flavour', None)
        if type(flavours) is not list:
            flavours = [flavours] * len(output)
            
        return isopy.refval({isopy.keystring(k, flavour=flavours[i]): v
                for i, (k,v) in enumerate(output.items())})
        
    elif output_type == 'ndarray':
        datatype = description.get('dtype', None)
        ndim = description.get('ndim', '2')
        if ndim == '0':
            output = output[0][0]
        elif ndim == '1':
            output = output[0]
        
        return np.array(output, dtype=datatype)

# TODO include unit description
# TODO preserve default value, store dtype as well
def data_to_rows(data, keys_in_first, keyfmt = None):
    data = isopy.asanyarray(data, flavour='general')
    
    if isinstance(data, core.IsopyArray):
        # TODO in to_list() have include columnkeys and rowID
        if data.ndim == 0:
            data = data.reshape(-1)
        if keys_in_first == 'r':
            rows = [[k.str(keyfmt) for k in data.keys()]]
            rows += data.to_list()
        elif keys_in_first == 'c':
            rows = [[k.str(keyfmt)] + v.tolist() for k, v in data.items()]
        else:
            raise ValueError(f'Unknown value for "keys_in_first" {keys_in_first}')
    else:
        if data.ndim == 0:
            data = data.reshape((1, 1))
        elif data.ndim == 1:
            data = data.reshape((1, -1))
        elif data.ndim > 2:
            raise ValueError('data cannot have more than 2 dimensions')
        rows = data.tolist()

    return rows

def data_description(data, keys_in_first):
    description = {}
    if isinstance(data, core.IsopyArray):
        description['keys_in_first'] = keys_in_first
        description['type'] = 'array'
        description['ndim'] = f'{data.ndim}'
        description['flavour'] = ";".join([str(f) for f in data.keys.flavours])
        description['dtype'] = ";".join([str(dt) for dt in data.datatypes])
    elif isinstance(data, core.RefValDict):
        raise NotImplementedError('RefValDicts cannot be saved with a description')
        description['type'] = 'refval'
        description['flavour'] = ";".join([str(f) for f in data.keys.flavours])
        if data.ratio_default is not None:
            description['ratio_default'] = data.ratio_default
        if data.molecule_default is not None:
            description['molecule_default'] = data.molecule_default
        # TODO default_value
    elif isinstance(data, np.ndarray):
        description['type'] = 'ndarray'
        description['ndim'] = f'{data.ndim}'
        description['dtype'] = str(data.dtype)
    elif isinstance(data, dict):
        description['keys_in_first'] = keys_in_first
    
    if description:
        return f'[isopy]{"&".join([f"{k}={v}" for k,v in description.items()])}'
    else:
        return None
    
def parse_description(string):
    description = {}
    for item in string.split('&'):
        k, v = item.split('=')
        if ';' in v:
            v = v.split(';')
        description[k] = v
    return description
    
################
### read exp ###
################
class NeptuneData:
    """
    Container for the data returned by ``read_exp``.
    """
    def __init__(self, info, cycle, time, measurements):
        self.info = info
        self.cycle = cycle
        self.time = time
        self.measurements = measurements


def read_exp(filename, rename = None) -> NeptuneData:
    """
    Load data from a Neptune/Triton export file.

    Parameters
    ----------
    filename : str, bytes, StringIO, BytesIO
        Path for file to be opened. Alternatively a file like byte string or a file like object can be supplied.
    rename : dict, Callable, Optional
        For renaming keys in the analysed data. Useful for cases when the key is the mass rather
        than the isotope measured. If a dictionary is passed then every key present in the
        dictionary will be replaced by the associated value. A callable can also be passed that
        takes the key in the file and returns the new key.

    Returns
    -------
    neptune_data : NeptuneData
        An object containing the following attributes:

        * info - Dictionary containing the metadata included at the beginning of the file.
        * cycle - A list containing the cycle number for each measurement.
        * time - A list containing datetime objects for each measurement.
        * measurements - An dictionary containing an an isopy array with the values in for each line measured. Static measurements are always given as line ``1``.

         To extract e.g only the isotope data from a measurement use ``neptune_data.measurement[line].copy(flavour_eq='isotope')``.
    """


    # If filename is a string load the files.
    if type(filename) is str:
        with open(filename, 'rb') as fileio:
            file = fileio.read()
    elif type(filename) is bytes:
        file = filename
    elif type(filename) is io.BytesIO:
        filename.seek(0)
        file = filename.read()
    elif type(filename) is io.StringIO:
        filename.seek(0)
        file = filename.read()
    else:
        raise TypeError('filename is of unknown type')

    if type(file) is bytes:
        # find the files encoding
        encoding = chardet.detect(file).get('encoding')

        # Decode the bytes into string.
        file = file.decode(encoding)

    csv_reader = csv.reader(io.StringIO(file), dialect='excel-tab')

    information = {}
    for row in csv_reader:
        if ':' in row[0]:  # meta
            name, value = row[0].split(':', 1)
            information[name.strip()] = value.strip()

        if row[0] == 'Cycle':
            # We can reuse the function from before to read the integration data
            data = _read_csv_data(row, csv_reader, termination_symbol='***')
            data = rows_to_data(data, True, 'r')

    if rename is None:
        renamer = lambda name: name
    elif isinstance(rename, dict):
        renamer = lambda name: rename.get(name, name)
    elif callable(rename):
        renamer = rename
    else:
        raise TypeError('rename must be a dict or a callable function')

    data.pop("", None)
    cycle = [int(i) for i in data.pop('Cycle')]

    time = data.pop('Time')
    try:
        time = [dt.datetime.strptime(time[i], '%H:%M:%S:%f') for i in range(len(time))]
    except:
        # Dont think the time will ever be given as a float
        # time = [dt.datetime.fromtimestamp(time[i]).strftime('%H:%M:%S:%f') for i in range(len(time))]
        warnings.warn('Unable to parse the time column')
        time = [dt.datetime.fromtimestamp(0).strftime('%H:%M:%S:%f') for i in range(len(time))]

    measurements = {}

    for key in data:
        # Check if the data has more than one line
        if ":" in key:
            line, keystring = key.split(":", 1)
            line = int(line)
        else:
            line = 1
            keystring = key

        # Rename columns if needed
        keystring = renamer(keystring)
        measurements.setdefault(line, {})[isopy.keystring(keystring)] = data[key]

    # Convert to isopy arrays
    measurements = {line: isopy.asarray(measurements[line]) for line in measurements}

    return NeptuneData(information, cycle, time, measurements)

######################
### read/write CSV ###
######################
def read_csv(filename, has_keys= None, keys_in_first = None, comment_symbol ='#', 
             encoding = None, dialect = 'excel', ignore_description = False):
    """
    Load data from a csv file.

    Parameters
    ----------
    filename : str, bytes, StringIO, BytesIO
        Path for file to be opened. Alternatively a file like byte string can be supplied.
        Also accepts file like objects.
    has_keys : bool, None
        If True or *keys_in_first* is not None a dictionary will always be returned. If False an nexted list is
        always returned. If None it will return a nested list of all values in the first row and column
        can be converted to/is a float.
    keys_in_first : {'c', 'r', None}
        Where the keys are found. Give 'r' if the keys are found in the first row and 'c' if the
        keys are in first column. If None it will analyse the data to determine where the keys are. If *has_keys*
        is not False an exception will be raised if it cant determine where the keys are.
    comment_symbol : str, Default = '#'
        Rows starting with this string will be ignored.
    encoding : str
        Encoding of the file. If None the encoding will be guessed from the file.
    dialect
        Dialect of the csv file. If None the dialect will be guessed from the file.
    ignore_description : bool
        If True the description of the array, if included in the file, is ignored.

    Returns
    -------
    data : dict or list
        Returns a dictionary for data with keys otherwise a list.
    """

    #If filename is a string load the files.
    if type(filename) is str:
        with open(filename, 'rb') as fileio:
            text = fileio.read()
    elif type(filename) is bytes:
        text = filename
    elif type(filename) is io.BytesIO:
        filename.seek(0)
        text = filename.read()
    elif type(filename) is io.StringIO:
        filename.seek(0)
        text = filename.read()
    else:
        raise TypeError(f'filename is of unknown type {type(filename)}')

    if type(text) is bytes:
        #find the files encoding
        if encoding is None:
            encoding = chardet.detect(text).get('encoding')

        #Decode the bytes into string.
        text = text.decode(encoding)

    #find the csv dialect
    if dialect is None:
        dialect = csv.Sniffer().sniff(text)

    description = None
    
    # Create a reader object by converting the file string to a file-like object
    csv_reader = csv.reader(io.StringIO(text), dialect=dialect)
    for row in csv_reader:
        row_data = [r.strip() for r in row]  # Remove any training whitespaces from the data in this row
        if len(row_data) == 0:
            return rows_to_data([[]], has_keys, keys_in_first)

        if comment_symbol is not None and row[0].startswith(comment_symbol):
            if not ignore_description and row[0].startswith(f'{comment_symbol}[isopy]'):
                description = parse_description(row[0].removeprefix(f'{comment_symbol}[isopy]'))
        else:
            data = _read_csv_data(row_data, csv_reader, comment_symbol)
            return rows_to_data(data, has_keys, keys_in_first, description)

    return rows_to_data([[]], has_keys, keys_in_first)



def _read_csv_data(first_row, reader, comment_symbol=None, termination_symbol=None):
    data = []
    dlen = None

    # Loop over the remaining rows
    for i, row in enumerate(itertools.chain([first_row], reader)):
        row = [r.strip() for r in row]
        if len(row) > 0:
            if termination_symbol is not None and row[0][:len(termination_symbol)] == termination_symbol:
                # Stop reading data if we find this string at the beginning of a row
                break

            if comment_symbol is not None and row[0][:len(comment_symbol)] == comment_symbol and row[0] not in NAN_STRINGS:
                # Row is a comment, ignore
                continue

        data.append([])
        if dlen is None:
            dlen = len(row)

        elif dlen != len(row):
            raise ValueError('Rows have different number of values')

        for j in range(len(row)):
            value = row[j]

            if value in NAN_STRINGS:
                value = float('nan')

            data[-1].append(value)

    return data


def write_csv(filename, data, comments=None, keys_in_first='r', comment_symbol = '#', keyfmt = None,
              dialect = 'excel', include_description=False) -> None:
    """
    Save data to a csv file.

    Parameters
    ----------
    filename : str, StringIO, BytesIO
        Path/name of the csv file to be created. Any existing file with the same path/name will be
        over written. Also accepts file like objects.
    data : isopy_array_like, numpy_array_like
        Data to be saved in the array.
    comments : str, Sequence[str], Optional
        Comments to be included at the top of the file
    keys_in_first : {'c', 'r'}
        Only used if the input has keys. Give 'r' if the keys should be in the first row and 'c' if the
        keys should be in the first column.
    comment_symbol : str, Default = '#'
        This string will precede any comments at the beginning of the file.
    keyfmt
        Specify the format used for the key string. See the ``str()`` method of each key string for options.
    dialect
        The CSV dialect used to save the file. Default to 'excel' which is a ', ' separated file.
    include_description : bool
        If True then a comment including a description of the data is included at the top of the file.
    """    
    if include_description:
        description = data_description(data, keys_in_first=keys_in_first)
    else:
        description = None
        
    rows = data_to_rows(data, keys_in_first, keyfmt)

    if comments is not None or description:
        if comments is None:
            comments = []   
        elif type(comments) is not list:
            comments = [comments]
        
        if description:
            comments = [description] + comments

        crows = []
        if comments is not None:
            for comment in comments:
                crows.append([f'{comment_symbol}{comment}'] + ['' for i in range(len(rows[0][1:]))])

        if len(rows[0]) == 0:
            rows = crows
        else:
            rows = crows + rows

    if type(filename) is str:
        with open(filename, mode='w', newline='') as file:
            csv_writer = csv.writer(file, dialect=dialect)
            for row in rows:
                csv_writer.writerow(row)

    elif type(filename) is io.StringIO:
        filename.truncate(0)
        filename.seek(0)
        csv_writer = csv.writer(filename, dialect=dialect)
        for row in rows:
            csv_writer.writerow(row)

    elif type(filename) is io.BytesIO:
        filename.truncate(0)
        filename.seek(0)

        file = io.StringIO()
        csv_writer = csv.writer(file, dialect=dialect)
        for row in rows:
            csv_writer.writerow(row)
        file.seek(0)
        text = file.read()
        filename.write(text.encode('UTF-8'))
    else:
        raise TypeError(f'filename is of unknown type {type(filename)}')

def read_clipboard(has_keys= None, keys_in_first = None, comment_symbol ='#', dialect = None):
    """
    Load data from values in the clipboard.

    Parameters
    ----------
    comment_symbol : str, Default = '#'
        Rows starting with this string will be ignored.
    has_keys : bool, None
        If True or *keys_in_first* is not None a dictionary will always be returned. If False an nexted list is
        always returned. If None it will return a nested list of all values in the first row and column
        can be converted to/is a float.
    keys_in_first : {'c', 'r', None}
        Where the keys are found. Give 'r' if the keys are found in the first row and 'c' if the
        keys are in first column. If None it will analyse the data to determine where the keys are. If *has_keys*
        is not False an exception will be raised if it cant determine where the keys are.
    dialect
        Dialect of the values in the clipboard. If None the dialect will be guessed from the values.
    """
    data = pyperclip.paste()
    return read_csv(io.StringIO(data), comment_symbol=comment_symbol, has_keys=has_keys,
                    keys_in_first=keys_in_first, dialect=dialect)

def write_clipboard(data, comments=None, keys_in_first='r',
              comment_symbol = '#', keyfmt = None, dialect = 'excel'):
    """
    Copies data to the clipboard

    Parameters
    ----------
    data : isopy_array_like, numpy_array_like
        Data to be copied.
    comments : str, Sequence[str], Optional
        Comments to be included
    keys_in_first : {'c', 'r'}
        Only used if the input has keys. Give 'r' if the keys should be in the first row and 'c' if the
        keys should be in the first column.
    comment_symbol : str, Default = '#'
        This string will precede any comments.
    keyfmt
        Specify the format used for the key string. See the ``str()`` method of each key string for options.
    dialect
        The CSV dialect used to copy the data to the clipboard. Default to 'excel' which is gives ', ' seperated data.
    """
    text = io.StringIO()
    write_csv(text, data, comments=comments, keys_in_first=keys_in_first, dialect=dialect,
              comment_symbol=comment_symbol, keyfmt=keyfmt)

    text.seek(0)
    pyperclip.copy(text.read())

#######################
### Read/write xlsx ###
#######################
def read_xlsx(filename, sheetname=None, has_keys = None, keys_in_first=None, comment_symbol='#', start_at='A1'):
    """
    Load data from an excel file.

    Parameters
    ----------
    filename : str, bytes, BytesIO
        Path for file to be opened. Also accepts file like objects.
    sheetname : str, int, Optional
        To load data from a single sheet in the workbook pass either the name of the sheet or
        the position of the sheet. If nothing is specified all the data for all sheets in the
        workbook is returned.
    has_keys : bool, None
        If True or *keys_in_first* is not None a dictionary will always be returned. If False an nexted list is
        always returned. If None it will return a nested list of all values in the first row and column
        can be converted to/is a float.
    keys_in_first : {'c', 'r', None}
        Where the keys are found. Give 'r' if the keys are found in the first row and 'c' if the
        keys are in first column. If None it will analyse the data to determine where the keys are. If *has_keys*
        is not False an exception will be raised if it cant determine where the keys are.
    comment_symbol : str, Default = '#'
        Rows starting with this string will be ignored.
    start_at : str, (int, int)
        Start scanning at this cell. Can either be a excel style cell reference or a (row, column) tuple of integers.

    Returns
    -------
    data : dict
        A dictionary containing the data for a single sheet or a dictionary containing the data
        for multiple sheets in the workbook indexed by sheet name.
    """
    if type(filename) is bytes:
        filename = io.BytesIO(filename)

    if type(filename) is openpyxl.Workbook:
        workbook = filename
    else:
        workbook = openpyxl.load_workbook(filename, data_only=True, read_only=True)

    sheetnames = workbook.sheetnames

    if sheetname is None:
        sheet_data = {}
        for sheetname in sheetnames:
            sheet_data[sheetname] = _read_xlsx_sheet(workbook[sheetname], has_keys, keys_in_first, comment_symbol, start_at)
        return sheet_data

    elif type(sheetname) is int and sheetname < len(sheetnames):
        return _read_xlsx_sheet(workbook[sheetnames[sheetname]], has_keys, keys_in_first, comment_symbol, start_at)

    elif sheetname in sheetnames:
        return _read_xlsx_sheet(workbook[sheetname], has_keys, keys_in_first, comment_symbol, start_at)

    else:
        raise ValueError(f'sheetname "{sheetname}" not found in workbook"')


def _read_xlsx_sheet(worksheet, has_keys, keys_in_first, comment_symbol, start_at):
    if type(start_at) is tuple:
        start_r, start_c = start_at
    else:
        start_r, start_c = openpyxl.utils.cell.coordinate_to_tuple(start_at)

    #Remove inital comments
    for ri in range(start_r, worksheet.max_row + 1):
        value = worksheet.cell(ri, start_c).value
        if type(value) is str and value[:len(comment_symbol)] == comment_symbol and value not in NAN_STRINGS:
            continue
        else:
            start_r = ri
            break

    # Select area containing data
    stop_c = worksheet.max_column + 1
    stop_r = worksheet.max_row + 1

    for ri in range(start_r + 1, stop_r+1):
        values_r = []
        for ci in range(start_c, stop_c):
            if ri != stop_r:
                # All rows should be included
                cell = worksheet.cell(ri, start_c)
                value = cell.value

                if value == '':
                    values_r.append(None)
                else:
                    values_r.append(value)

            if values_r.count(None) == len(values_r):
                values_c = []
                if ci != (stop_c-1):
                    for rj in range(start_r, ri + 1):
                        cell = worksheet.cell(rj, ci + 1)
                        value = cell.value

                        if value == '':
                            values_c.append(None)
                        else:
                            values_c.append(value)

                if values_c.count(None) == len(values_c):
                    stop_r = ri
                    stop_c = ci + 1

            if ci == stop_c: break
        if ri == stop_r: break

    if (stop_r-start_r) == 1 and (stop_c-start_c) == 1 and ((cell:=worksheet.cell(start_r,start_c)).value is None or cell.value==''):
        return rows_to_data([[]], has_keys, keys_in_first)

    data = []

    # Iterate over each column
    for ri in range(start_r, stop_r):
        value = worksheet.cell(ri, start_c).value
        if type(value) is str and value[:len(comment_symbol)] == comment_symbol and value not in NAN_STRINGS:
            continue

        data.append([])
        for ci in range(start_c, stop_c):
            cell = worksheet.cell(ri, ci)
            value = cell.value
            if type(value) is str:
                value = value.strip()

            if cell.data_type == 'e' or (type(value) is str and value.upper() == '=NA()'):
                data[-1].append(float('nan'))
            elif value == '' or value in NAN_STRINGS or value is None:
                data[-1].append(float('nan'))
            else:
                data[-1].append(value)

    return rows_to_data(data, has_keys, keys_in_first)


def write_xlsx(filename, *sheets, comments = None,
               keys_in_first= 'r', comment_symbol= '#', keyfmt = None,
               start_at ="A1", append = False, clear = True, **sheetnames):
    """
    Save data to an excel file.

    Parameters
    ----------
    filename : str, BytesIO
        Path/name of the excel file to be created. Any existing file with the same path/name
        will be overwritten. Also accepts file like objects.
    sheets : isopy_array_like, numpy_array_like
        Data given here will be saved as sheet1, sheet2 etc.
    comments : str, Sequence[str], Optional
        Comments to be included at the top of the file
    keys_in_first : {'c', 'r'}
        Only used if the input has keys. Give 'r' if the keys should be in the first row and 'c' if the
        keys should be in the first column.
    comment_symbol : str, Default = '#'
        This string will precede any comments at the beginning of the file
    keyfmt
        Specify the format used for the key string. See the ``str()`` method of each key string for options.
    start_at: str, (int, int)
        The first cell where the data is written. Can either be a excel style cell reference or a (row, column)
        tuple of integers.
    append : bool, Default = False
        If ``True`` and *filename* exists it will append the data to this workbook. An exception
        is raised if *filename* is not a valid excel workbook.
    clear : bool, Default = True
        If ``True`` any preexisting sheets are cleared before any new data is written to it.
    sheetnames : isopy_array, numpy_array
        Data given here will be saved in a sheet with *sheetname* name.
    """
    save = True
    if type(filename) is openpyxl.Workbook:
        workbook = filename
        save = False
    elif type(filename) is io.BytesIO:
        if append and filename.seek(0,2) > 0:
            workbook = openpyxl.load_workbook(filename=filename)
        else:
            # openpyxl truncates BytesIO objects automatically via the use of ZipFile('w')

            workbook = openpyxl.Workbook()
            workbook.remove(workbook.active)
    elif type(filename) is str:
        if append and _os.path.exists(filename):
            workbook = openpyxl.load_workbook(filename=filename)
        else:
            workbook = openpyxl.Workbook()
            workbook.remove(workbook.active)

    sheetname_data = {f'sheet{i+1}': data for i, data in enumerate(sheets)}
    sheetname_data.update(sheetnames)
    try:
        for sheetname, data in sheetname_data.items():
            #If appending then delete any preexisting workbooks
            if sheetname not in workbook.sheetnames:
                worksheet = workbook.create_sheet(sheetname)
            elif clear:
                workbook.remove(workbook[sheetname])
                worksheet = workbook.create_sheet(sheetname)
            else:
               worksheet = workbook[sheetname]

            _write_xlsx(worksheet, data, comments, comment_symbol, keys_in_first, start_at, keyfmt)

        #Workbooks must have at least one sheet
        if len(workbook.sheetnames) == 0:
            workbook.create_sheet('sheet1')

        if save: workbook.save(filename)
    finally:
        if save: workbook.close()

def _write_xlsx(worksheet, data, comments, comment_symbol, keys_in_first, start_at, keyfmt):
    rows = data_to_rows(data, keys_in_first, keyfmt)
    if type(start_at) is tuple:
        start_r, start_c = start_at
    else:
        start_r, start_c = openpyxl.utils.cell.coordinate_to_tuple(start_at)

    if comments is not None:
        if not isinstance(comments, (list, tuple)):
            comments = [comments]
        for comment in comments:
            worksheet.cell(start_r, start_c).value = f'{comment_symbol}{comment}'
            start_r += 1

    for ri, row in enumerate(rows):
        for ci, value in enumerate(row):
            if str(value) == 'nan':
                value = '#N/A'
            if value is not None and value != '':
                worksheet.cell(start_r + ri, start_c + ci).value = value

# Archive
# TODO what happens to groups without any files in them
# __repr__, __str__ and _repr_markdown_ 

def verify_archive_filename(filename, overwrite = True, allow_none = False):
    if type(filename) is str:
        if not filename.endswith('.zip') and not filename.endswith('.data'):
            filename = filename + '.data.zip'
        if not overwrite and os.path.exists(filename):
            raise ValueError(f'"{filename}" already exists')
    elif allow_none and filename is None:
        pass
    elif type(filename) is not io.BytesIO:
        raise TypeError(f'Filename must be a string or a BytesIO object (not {type(filename)})')
    
    return filename

def new_archive(filename = None, overwrite=False):
    """ Create a new data archive

    Parameters
    ----------
    filename : str, io.BytesIO, optional
        The filename, or file like object, where the archive will be saved.
    overwrite : bool, optional
        If False an exception is raised if *filename* already exits.

    Returns
    -------
    DataArchive
    """
    filename = verify_archive_filename(filename, overwrite=overwrite, allow_none=True)
    return DataArchive(filename)
    

def load_archive(filename):
    """ Loads a data archive

    Parameters
    ----------
    filename : str, io.BytesIO
        The filename, or file like object, where the archive is saved.

    Returns
    -------
    DataArchive

    Raises
    ------
    ValueError
        Raised if *filename* does not exist.
    """
    filename = verify_archive_filename(filename)
    
    archive = DataArchive(filename)
    archive.reload()
    return archive

class DataGroup:
    def __init__(self):
        super().__setattr__('_items', {})
        
    def _repr_markdown_(self):
        return self.__str__()
    
    def __str__(self):
        return f'This archive group contains:{self._str_()}'
    
    def _str_(self, level = 0):
        string = ''
        pre = f'\n{" "*(level*2)}- '
        for k, v in self._items.items():
            if isinstance(v, DataGroup):
                string += f'{pre}{k}/'
                string += v._str_(level+1)
            elif isinstance(v, core.IsopyArray):
                string += f'{pre}{k} (ndim={v.ndim}, nrows={v.nrows}, ncols={v.ncols})'
            else:
                string += f'{pre}{k} (ndim={v.ndim}, size={v.size})'
        return string
    
    def __repr__(self):
        return 'DataGroup([' + self._repr_().removeprefix(', ') + '])'
        
    def _repr_(self, path = ''):
        string = ''
        for k, v in self._items.items():
            if isinstance(v, DataGroup):
                string += f', {path}{k}/'
                string += v._repr_(f'{path}{k}/')
            else:
                string += f', {path}{k}'
        return string
    
    def _save_(self, file, path):
        for name, data in self._items.items():
            if isinstance(data, DataGroup):
                data._save_(file, f'{path}{name}/')
            else:
                isopy.write_csv((b:=io.BytesIO()), data, include_description=True)
                file.writestr(f'{path}{name}.csv', b.getvalue())
    
    def __getattr__(self, name):
        try:
            return self._items[name]
        except KeyError:
            raise AttributeError(f'No attribute, group or data with name "{name}"')
    
    def __setattr__(self, name, data):
        self._add_(name.split('__'), data)
        
    def __delattr__(self, name):
        item = self._items.pop(name, None)
        if isinstance(item, DataGroup):
            item.clear()
    
    def _add_(self, names, data):
        if len(names) == 1 and data is not None:
            name = names[0]
            if not name.isidentifier():
                raise ValueError(f'Invalid name: {name}')

            data = isopy.asanyarray(data)
            
            if name in self._items and isinstance(self._items[name], DataGroup):
                raise ValueError('f"{name}" is being used by a group item.')
            else:
                self._items[name] = data
        elif len(names) > 0:
            name = names[0]
            if not name.isidentifier():
                raise ValueError(f'Invalid name: {name}')
            
            if name not in self._items:
                self._items[name] = DataGroup()
            elif name in self._items and not isinstance(self._items[name], DataGroup):
                raise ValueError('f"{name}" is being used by a data item.')
            
            return self._items[name]._add_(names[1:], data)
        
        else:
            return self
                
    def add(self, name, data = None):
        """Add subgroup or data to the current location in the archive.

        Parameters
        ----------
        name : str
            Name or path of the subgroup or data relative to the current location within the archive. Can also be a path within the archive. 
            Any non-existing groups will be created as necessary. Existing subgroups will not be overwritten but existing data will be.
            
        data : array, ndarray, optional
            The data to be included in the archive. Argument can be anything that is compatible with `isopy.asanyarray`.
            If not given then a subgroup is created instead.

        Returns
        -------
        array, DataGroup
            Returns the subgroup or data added to the archive.
            
        Raises
        ------
        ValueError
            If *name* is not a valid python identifier.
        """
        paths = name.split('/')
        for name in paths:
            if not name.isidentifier():
                raise ValueError(f'Invalid name: {name}')
        
        return self._add_(paths, data)
        
    def clear(self):
        """Removes all data and subgroups at the current location within the archive. 
        """
        for item in self._items.values():
            if isinstance(item, DataGroup):
                item.clear()
            
        self._items.clear()
    

class DataArchive(DataGroup):
    """ A data archive that can contain multiple isopy arrays, refvals and 1D/2D numpy arrays.
    
    Subgroups and data can be accessed as attributes, e.g. `archive.data1` or `archive.group1.data1`.

    Attributes
    ----------
    filename : str, io.BytesIO, None
        The filename, or file like object, where the archive is saved. Readonly
    """
    def __init__(self, filename = None):
        super().__init__()
        super(DataGroup, self).__setattr__('_filename', filename)
    
    def __str__(self):
        return f'The archive {self.filename} contains:{self._str_()}'
    
    def __repr__(self):  
        return f'DataArchive({repr(self.filename)}, [' + self._repr_().removeprefix(', ') + '])'
    
    @property
    def filename(self):
        return self._filename
    
    def _loadfile_(self, filename):
        self.clear()
        
        if filename is None:
            return
        
        if type(filename) is io.BytesIO:
            filename.seek(0)
        
        with zipfile.ZipFile(filename, mode='r') as file:
            for path in file.namelist():
                # Only load csv files, other files ignored
                if path.endswith('.csv'):
                    data = isopy.read_csv(file.read(path))
                    self._add_(path.removesuffix('.csv').split('/'), data)
        
    def _savefile_(self, filename):
        # Automatically truncates BytesIO objects       
        with zipfile.ZipFile(filename, mode='w') as file:
            self._save_(file, '')
        
    def save(self):
        """ Save the archive.
        
        This method is not available for archive subgroups.
        """
        if self._filename is None:
            raise ValueError('No filename or file object associated with the Archive')
        
        self._savefile_(self._filename)
    
    def saveas(self, filename, overwrite=False):
        """ Save the archive to *filename*.
        
        Future calls to the `save()` method will save the archive to *filename*.
        
        This method is not available for archive subgroups.

        Parameters
        ----------
        filename : str, io.BytesIO
            The new location of the archive.
        overwrite : bool, optional
            Whether existing files should be overwritten.
            
        Raises
        ------
        IOError
            Raised if *filename* already exits and *overwrite* is False.
        """
        filename = verify_archive_filename(filename, overwrite)
        self._savefile_(filename)
        super(DataGroup, self).__setattr__('_filename', filename)
    
    def reload(self):
        """Reloads the contents of the archive. Any unsaved changes are lost.
        
        This method is not available for archive subgroups.
        """
        if self._filename is None:
            raise ValueError('No filename or file object associated with the Archive')
        
        self._loadfile_(self._filename)
