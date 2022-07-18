import datetime

import pyperclip

import isopy
import numpy as np
import os
import time
import io
import pytest
import csv
import warnings

import isopy.core as core
import isopy.io

import openpyxl

sleep_time = 0.1

def filename(filename):
    return os.path.join(os.path.dirname(__file__), 'files', filename)

class Test_Data_rows:
    def test_data_to_rows_3_1(self):
        data_d = dict(Pd=[1.1, 2.2, 3.3], Cd=[11.1, np.nan, 13.3], Ru=[21.1, 22.2, 23.3])
        data_ds = core.RefValDict(data_d)
        data_a = isopy.array(data_d)
        data_l = data_a.to_list()
        data_nd = np.array(data_l)

        # keys in first row
        true = [list(data_d.keys())] + data_l

        data = isopy.io.data_to_rows(data_d, 'r')
        self.compare_rows(data, true, 'r')

        data = isopy.io.data_to_rows(data_a, 'r')
        self.compare_rows(data, true, 'r')

        data = isopy.io.data_to_rows(data_ds, 'r')
        self.compare_rows(data, true, 'r')

        # keys in first column
        true = [[k] + v for k, v in data_d.items()]

        data = isopy.io.data_to_rows(data_d, 'c')
        self.compare_rows(data, true, 'c')

        data = isopy.io.data_to_rows(data_a, 'c')
        self.compare_rows(data, true, 'c')

        data = isopy.io.data_to_rows(data_ds, 'c')
        self.compare_rows(data, true, 'c')

        # No keys
        true = data_l

        data = isopy.io.data_to_rows(data_l, 'r')
        self.compare_rows(data, true)

        data = isopy.io.data_to_rows(data_nd, 'r')
        self.compare_rows(data, true)

        data = isopy.io.data_to_rows(data_l, 'c')
        self.compare_rows(data, true)

        data = isopy.io.data_to_rows(data_nd, 'c')
        self.compare_rows(data, true)

    def test_data_to_rows_1_1(self):
        data_d = dict(Pd=[1.1], Cd=[11.1], Ru=[21.1])
        data_ds = core.RefValDict(data_d)
        data_a = isopy.array(data_d)
        data_l = data_a.to_list()
        data_nd = np.array(data_l)

        # keys in first row
        true = [list(data_d.keys())] + data_l

        data = isopy.io.data_to_rows(data_d, 'r')
        self.compare_rows(data, true, 'r')

        data = isopy.io.data_to_rows(data_a, 'r')
        self.compare_rows(data, true, 'r')

        data = isopy.io.data_to_rows(data_ds, 'r')
        self.compare_rows(data, true, 'r')

        # keys in first column
        true = [[k] + v for k, v in data_d.items()]

        data = isopy.io.data_to_rows(data_d, 'c')
        self.compare_rows(data, true, 'c')

        data = isopy.io.data_to_rows(data_a, 'c')
        self.compare_rows(data, true, 'c')

        data = isopy.io.data_to_rows(data_ds, 'c')
        self.compare_rows(data, true, 'c')

        # No keys
        true = data_l

        data = isopy.io.data_to_rows(data_l, 'r')
        self.compare_rows(data, true)

        data = isopy.io.data_to_rows(data_nd, 'r')
        self.compare_rows(data, true)

        data = isopy.io.data_to_rows(data_l, 'c')
        self.compare_rows(data, true)

        data = isopy.io.data_to_rows(data_nd, 'c')
        self.compare_rows(data, true)

    def test_data_to_rows_1_0(self):
        data_d = dict(Pd=1.1, Cd=11.1, Ru=21.1)
        data_ds = core.RefValDict(data_d)
        data_a = isopy.array(data_d)
        data_l = data_a.to_list()
        data_nd = np.array(data_l)

        # keys in first row
        true = [list(data_d.keys())] + [data_l]

        data = isopy.io.data_to_rows(data_d, 'r')
        self.compare_rows(data, true, 'r')

        data = isopy.io.data_to_rows(data_a, 'r')
        self.compare_rows(data, true, 'r')

        data = isopy.io.data_to_rows(data_ds, 'r')
        self.compare_rows(data, true, 'r')

        # keys in first column
        true = [[k] + [v] for k, v in data_d.items()]

        data = isopy.io.data_to_rows(data_d, 'c')
        self.compare_rows(data, true, 'c')

        data = isopy.io.data_to_rows(data_a, 'c')
        self.compare_rows(data, true, 'c')

        data = isopy.io.data_to_rows(data_ds, 'c')
        self.compare_rows(data, true, 'c')

        # No keys
        true = [data_l]

        data = isopy.io.data_to_rows(data_l, 'r')
        self.compare_rows(data, true)

        data = isopy.io.data_to_rows(data_nd, 'r')
        self.compare_rows(data, true)

        data = isopy.io.data_to_rows(data_l, 'c')
        self.compare_rows(data, true)

        data = isopy.io.data_to_rows(data_nd, 'c')
        self.compare_rows(data, true)

        # No keys
        data_l = 11.1
        data_nd = np.array(data_l)
        true = [[data_l]]

        data = isopy.io.data_to_rows(data_l, 'r')
        self.compare_rows(data, true)

        data = isopy.io.data_to_rows(data_nd, 'r')
        self.compare_rows(data, true)

        data = isopy.io.data_to_rows(data_l, 'c')
        self.compare_rows(data, true)

        data = isopy.io.data_to_rows(data_nd, 'c')
        self.compare_rows(data, true)

    def test_rows_to_data_3_1_3(self):
        data_d = dict(Pd=[1.1, 2.2, 3.3], Cd=[11.1, np.nan, 13.3], Ru=[21.1, 22.2, 23.3])
        data_a = isopy.array(data_d)
        data_l = data_a.to_list()

        # keys in first row
        rows = [list(data_d.keys())] + data_l

        data = isopy.io.rows_to_data(rows, None, None)
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'r')
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'c')
        assert type(data) is dict
        with pytest.raises(AssertionError):
            self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, False, 'r')
        assert type(data) is not dict

        # Everything is a str
        rows = [[str(v) for v in row] for row in rows]

        data = isopy.io.rows_to_data(rows, None, None)
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'r')
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'c')
        assert type(data) is dict
        with pytest.raises(AssertionError):
            self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, False, 'r')
        assert type(data) is not dict

        # keys in first column
        rows = [[k] + v for k, v in data_d.items()]

        data = isopy.io.rows_to_data(rows, None, None)
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'c')
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'r')
        assert type(data) is dict
        with pytest.raises(AssertionError):
            self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, False, 'c')
        assert type(data) is not dict

        # Everything is a str
        rows = [[str(v) for v in row] for row in rows]

        data = isopy.io.rows_to_data(rows, None, None)
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'c')
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'r')
        assert type(data) is dict
        with pytest.raises(AssertionError):
            self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, False, 'c')
        assert type(data) is not dict

        # No keys
        rows = data_l

        data = isopy.io.rows_to_data(rows, None, None)
        self.compare_data(data, data_l, False)

        data = isopy.io.rows_to_data(rows, False, 'c')
        self.compare_data(data, data_l, False)

        data = isopy.io.rows_to_data(rows, False, 'r')
        self.compare_data(data, data_l, False)

        data = isopy.io.rows_to_data(rows, None, 'c')
        assert type(data) is dict

        data = isopy.io.rows_to_data(rows, None, 'r')
        assert type(data) is dict

        data = isopy.io.rows_to_data(rows, True, 'c')
        assert type(data) is dict

        data = isopy.io.rows_to_data(rows, True, 'r')
        assert type(data) is dict

        # Everything is a str
        rows = [[str(v) for v in row] for row in rows]

        data = isopy.io.rows_to_data(rows, None, None)
        self.compare_data(data, data_l, False)

        data = isopy.io.rows_to_data(rows, False, 'c')
        self.compare_data(data, data_l, False)

        data = isopy.io.rows_to_data(rows, False, 'r')
        self.compare_data(data, data_l, False)

        data = isopy.io.rows_to_data(rows, None, 'c')
        assert type(data) is dict

        data = isopy.io.rows_to_data(rows, None, 'r')
        assert type(data) is dict

        data = isopy.io.rows_to_data(rows, True, 'c')
        assert type(data) is dict

        data = isopy.io.rows_to_data(rows, True, 'r')
        assert type(data) is dict

    def test_rows_to_data_3_1_1(self):
        data_d = dict(Pd=[1.1, np.nan, 3.3])
        data_a = isopy.array(data_d)
        data_l = data_a.to_list()

        # keys in first row
        rows = [list(data_d.keys())] + data_l

        data = isopy.io.rows_to_data(rows, None, None)
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'r')
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'c')
        assert type(data) is dict
        with pytest.raises(AssertionError):
            self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, False, 'r')
        assert type(data) is not dict

        # Everything is a str
        rows = [[str(v) for v in row] for row in rows]

        data = isopy.io.rows_to_data(rows, None, None)
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'r')
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'c')
        assert type(data) is dict
        with pytest.raises(AssertionError):
            self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, False, 'r')
        assert type(data) is not dict

        # keys in first column
        rows = [[k] + v for k, v in data_d.items()]

        data = isopy.io.rows_to_data(rows, None, None)
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'c')
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'r')
        assert type(data) is dict
        with pytest.raises(AssertionError):
            self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, False, 'c')
        assert type(data) is not dict

        # Everything is a str
        rows = [[str(v) for v in row] for row in rows]

        data = isopy.io.rows_to_data(rows, None, None)
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'c')
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'r')
        assert type(data) is dict
        with pytest.raises(AssertionError):
            self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, False, 'c')
        assert type(data) is not dict

        # No keys
        rows = data_l

        data = isopy.io.rows_to_data(rows, None, None)
        self.compare_data(data, data_l, False)

        data = isopy.io.rows_to_data(rows, False, 'c')
        self.compare_data(data, data_l, False)

        data = isopy.io.rows_to_data(rows, False, 'r')
        self.compare_data(data, data_l, False)

        data = isopy.io.rows_to_data(rows, None, 'c')
        assert type(data) is dict

        data = isopy.io.rows_to_data(rows, None, 'r')
        assert type(data) is dict

        data = isopy.io.rows_to_data(rows, True, 'c')
        assert type(data) is dict

        data = isopy.io.rows_to_data(rows, True, 'r')
        assert type(data) is dict

        # Everything is a str
        rows = [[str(v) for v in row] for row in rows]

        data = isopy.io.rows_to_data(rows, None, None)
        self.compare_data(data, data_l, False)

        data = isopy.io.rows_to_data(rows, False, 'c')
        self.compare_data(data, data_l, False)

        data = isopy.io.rows_to_data(rows, False, 'r')
        self.compare_data(data, data_l, False)

        data = isopy.io.rows_to_data(rows, None, 'c')
        assert type(data) is dict

        data = isopy.io.rows_to_data(rows, None, 'r')
        assert type(data) is dict

        data = isopy.io.rows_to_data(rows, True, 'c')
        assert type(data) is dict

        data = isopy.io.rows_to_data(rows, True, 'r')
        assert type(data) is dict

    def test_rows_to_data_1_1(self):
        data_d = dict(Pd=[1.1], Cd=[11.1], Ru=[21.1])
        data_a = isopy.array(data_d)
        data_l = data_a.to_list()

        # keys in first row
        rows = [list(data_d.keys())] + data_l

        data = isopy.io.rows_to_data(rows, None, None)
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'r')
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'c')
        assert type(data) is dict
        with pytest.raises(AssertionError):
            self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, False, 'r')
        assert type(data) is not dict

        # Everything is a str
        rows = [[str(v) for v in row] for row in rows]

        data = isopy.io.rows_to_data(rows, None, None)
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'r')
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'c')
        assert type(data) is dict
        with pytest.raises(AssertionError):
            self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, False, 'r')
        assert type(data) is not dict

        # keys in first column
        rows = [[k] + v for k, v in data_d.items()]

        data = isopy.io.rows_to_data(rows, None, None)
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'c')
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'r')
        assert type(data) is dict
        with pytest.raises(AssertionError):
            self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, False, 'c')
        assert type(data) is not dict

        # Everything is a str
        rows = [[str(v) for v in row] for row in rows]

        data = isopy.io.rows_to_data(rows, None, None)
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'c')
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'r')
        assert type(data) is dict
        with pytest.raises(AssertionError):
            self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, False, 'c')
        assert type(data) is not dict

        # No keys
        rows = data_l

        data = isopy.io.rows_to_data(rows, None, None)
        self.compare_data(data, data_l, False)

        data = isopy.io.rows_to_data(rows, False, 'c')
        self.compare_data(data, data_l, False)

        data = isopy.io.rows_to_data(rows, False, 'r')
        self.compare_data(data, data_l, False)

        data = isopy.io.rows_to_data(rows, None, 'c')
        assert type(data) is dict

        data = isopy.io.rows_to_data(rows, None, 'r')
        assert type(data) is dict

        data = isopy.io.rows_to_data(rows, True, 'c')
        assert type(data) is dict

        data = isopy.io.rows_to_data(rows, True, 'r')
        assert type(data) is dict

        # Everything is a str
        rows = [[str(v) for v in row] for row in rows]

        data = isopy.io.rows_to_data(rows, None, None)
        self.compare_data(data, data_l, False)

        data = isopy.io.rows_to_data(rows, False, 'c')
        self.compare_data(data, data_l, False)

        data = isopy.io.rows_to_data(rows, False, 'r')
        self.compare_data(data, data_l, False)

        data = isopy.io.rows_to_data(rows, None, 'c')
        assert type(data) is dict

        data = isopy.io.rows_to_data(rows, None, 'r')
        assert type(data) is dict

        data = isopy.io.rows_to_data(rows, True, 'c')
        assert type(data) is dict

        data = isopy.io.rows_to_data(rows, True, 'r')
        assert type(data) is dict

    def test_rows_to_data_1_0_3(self):
        data_d = dict(Pd=1.1, Cd=11.1, Ru=21.1)
        data_a = isopy.array(data_d)
        data_l = data_a.to_list()

        # keys in first row
        rows = [list(data_d.keys())] + [data_l]

        data = isopy.io.rows_to_data(rows, None, None)
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'r')
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'c')
        assert type(data) is dict
        with pytest.raises(AssertionError):
            self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, False, 'r')
        assert type(data) is not dict

        # Everything is a str
        rows = [[str(v) for v in row] for row in rows]

        data = isopy.io.rows_to_data(rows, None, None)
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'r')
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'c')
        assert type(data) is dict
        with pytest.raises(AssertionError):
            self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, False, 'r')
        assert type(data) is not dict

        # keys in first column
        rows = [[k] + [v] for k, v in data_d.items()]

        data = isopy.io.rows_to_data(rows, None, None)
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'c')
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'r')
        assert type(data) is dict
        with pytest.raises(AssertionError):
            self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, False, 'c')
        assert type(data) is not dict

        # Everything is a str
        rows = [[str(v) for v in row] for row in rows]

        data = isopy.io.rows_to_data(rows, None, None)
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'c')
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'r')
        assert type(data) is dict
        with pytest.raises(AssertionError):
            self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, False, 'c')
        assert type(data) is not dict

        # No keys
        rows = [data_l]

        data = isopy.io.rows_to_data(rows, None, None)
        self.compare_data(data, [data_l], False)

        data = isopy.io.rows_to_data(rows, False, 'c')
        self.compare_data(data, [data_l], False)

        data = isopy.io.rows_to_data(rows, False, 'r')
        self.compare_data(data, [data_l], False)

        data = isopy.io.rows_to_data(rows, None, 'c')
        assert type(data) is dict

        data = isopy.io.rows_to_data(rows, None, 'r')
        assert type(data) is dict

        data = isopy.io.rows_to_data(rows, True, 'c')
        assert type(data) is dict

        data = isopy.io.rows_to_data(rows, True, 'r')
        assert type(data) is dict

        # Everything is a str
        rows = [[str(v) for v in row] for row in rows]

        data = isopy.io.rows_to_data(rows, None, None)
        self.compare_data(data, [data_l], False)

        data = isopy.io.rows_to_data(rows, False, 'c')
        self.compare_data(data, [data_l], False)

        data = isopy.io.rows_to_data(rows, False, 'r')
        self.compare_data(data, [data_l], False)

        data = isopy.io.rows_to_data(rows, None, 'c')
        assert type(data) is dict

        data = isopy.io.rows_to_data(rows, None, 'r')
        assert type(data) is dict

        data = isopy.io.rows_to_data(rows, True, 'c')
        assert type(data) is dict

        data = isopy.io.rows_to_data(rows, True, 'r')
        assert type(data) is dict

    def test_rows_to_data_1_0_1(self):
        data_d = dict(Cd=11.1)
        data_a = isopy.array(data_d)
        data_l = 11.1

        # keys in first row
        rows = [list(data_d.keys())] + [[data_l]]

        data = isopy.io.rows_to_data(rows, None, None)
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'r')
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'c')
        assert type(data) is dict
        with pytest.raises(AssertionError):
            self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, False, 'r')
        assert type(data) is not dict

        # Everything is a str
        rows = [[str(v) for v in row] for row in rows]

        data = isopy.io.rows_to_data(rows, None, None)
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'r')
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'c')
        assert type(data) is dict
        with pytest.raises(AssertionError):
            self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, False, 'r')
        assert type(data) is not dict

        # keys in first column
        rows = [[k] + [v] for k, v in data_d.items()]

        data = isopy.io.rows_to_data(rows, None, None)
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'c')
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'r')
        assert type(data) is dict
        with pytest.raises(AssertionError):
            self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, False, 'c')
        assert type(data) is not dict

        # Everything is a str
        rows = [[str(v) for v in row] for row in rows]

        data = isopy.io.rows_to_data(rows, None, None)
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'c')
        self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, None, 'r')
        assert type(data) is dict
        with pytest.raises(AssertionError):
            self.compare_data(data, data_d, True)

        data = isopy.io.rows_to_data(rows, False, 'c')
        assert type(data) is not dict

        # No keys
        rows = [[data_l]]

        data = isopy.io.rows_to_data(rows, None, None)
        self.compare_data(data, [[data_l]], False)

        data = isopy.io.rows_to_data(rows, False, 'c')
        self.compare_data(data, [[data_l]], False)

        data = isopy.io.rows_to_data(rows, False, 'r')
        self.compare_data(data, [[data_l]], False)

        data = isopy.io.rows_to_data(rows, None, 'c')
        assert type(data) is dict

        data = isopy.io.rows_to_data(rows, None, 'r')
        assert type(data) is dict

        data = isopy.io.rows_to_data(rows, True, 'c')
        assert type(data) is dict

        data = isopy.io.rows_to_data(rows, True, 'r')
        assert type(data) is dict

        # Everything is a str
        rows = [[str(v) for v in row] for row in rows]

        data = isopy.io.rows_to_data(rows, None, None)
        self.compare_data(data, [[data_l]], False)

        data = isopy.io.rows_to_data(rows, False, 'c')
        self.compare_data(data, [[data_l]], False)

        data = isopy.io.rows_to_data(rows, False, 'r')
        self.compare_data(data, [[data_l]], False)

        data = isopy.io.rows_to_data(rows, None, 'c')
        assert type(data) is dict

        data = isopy.io.rows_to_data(rows, None, 'r')
        assert type(data) is dict

        data = isopy.io.rows_to_data(rows, True, 'c')
        assert type(data) is dict

        data = isopy.io.rows_to_data(rows, True, 'r')
        assert type(data) is dict

    def test_rows_to_data_0(self):
        data = isopy.io.rows_to_data([[]], None, None)
        assert data == [[]]

        data = isopy.io.rows_to_data([[]], False, None)
        assert data == [[]]

        data = isopy.io.rows_to_data([[]], None, 'c')
        assert data == dict()

        data = isopy.io.rows_to_data([[]], False, 'r')
        assert data == dict()

        data = isopy.io.rows_to_data([[]], True, None)
        assert data == dict()

    def test_unclear_keys(self):
        # Cant determine which row has the keys
        rows = [['a', 'b', 'c'], ['d', 1, 2], ['e', 3, 4]]

        data = isopy.io.rows_to_data(rows, None, None)
        assert type(data) is dict
        assert list(data.keys()) == 'a b c'.split()

        data = isopy.io.rows_to_data(rows, True, None)
        assert type(data) is dict
        assert list(data.keys()) == 'a b c'.split()

        data = isopy.io.rows_to_data(rows, False, None)
        assert data is rows

        data = isopy.io.rows_to_data(rows, None, 'r')
        assert type(data) is dict
        assert list(data.keys()) == 'a b c'.split()

        data = isopy.io.rows_to_data(rows, None, 'c')
        assert type(data) is dict
        assert list(data.keys()) == 'a d e'.split()

        rows = [['a', 1, 2], [3, 4, 5], [6, 7, 7]]

        data = isopy.io.rows_to_data(rows, None, None)
        assert data is rows

        rows = [[str(r) for r in row] for row in rows]
        data = isopy.io.rows_to_data(rows, None, None)
        assert type(data) is dict
        assert list(data.keys()) == 'a 1 2'.split()

    def test_exceptions(self):
        rows = [['a', 'b', 'c'], ['d', 1, 2], ['e', 3, 4]]

        # Bad keys_in_first values
        with pytest.raises(ValueError):
            isopy.io.rows_to_data(rows, None, 'd')

        data = dict(Pd=[1.1, 2.2, 3.3], Cd=[11.1, np.nan, 13.3], Ru=[21.1, 22.2, 23.3])

        with pytest.raises(ValueError):
            isopy.io.data_to_rows(data, 'd')

        # Data has to many dimensions
        data = [[[1]], [[2]], [[3]]]

        with pytest.raises(ValueError):
            isopy.io.data_to_rows(data, 'c')

    def compare_rows(self, data, true, keys_in_first=None):
        if keys_in_first == 'r':
            assert data[0] == true[0]
            data = data[1:]
            true = true[1:]
        elif keys_in_first == 'c':
            assert [d[0] for d in data] == [t[0] for t in true]
            data = [d[1:] for d in data]
            true = [t[1:] for t in true]

        data = np.array(data, dtype=np.float64)
        true = np.array(true, dtype=np.float64)

        assert data.size == true.size
        assert data.ndim == 2

        np.testing.assert_allclose(data, true)

    def compare_data(self, data, true, has_keys):
        if has_keys:
            assert type(data) is dict
            assert list(data.keys()) == list(true.keys())

            for key in data.keys():
                d = np.array(data[key], dtype=np.float64)
                t = np.array(true[key], dtype=np.float64)

                assert d.size == t.size
                assert d.ndim == 1

                np.testing.assert_allclose(d, t)
        else:
            assert type(data) is list
            data = np.array(data, dtype=np.float64)
            true = np.array(true, dtype=np.float64)

            assert data.size == true.size
            assert data.ndim == 2

            np.testing.assert_allclose(data, true)

class Test_CSV:
    def test_all(self):
        # This fails occasionaly because it says the file is not avaliable
        # A short sleep between writing and reading appears to solve the problem
        data = dict(Pd=[1.0, 2.0, 3.0], Cd=[11.0, np.nan, 13.0], Ru=[21.0, 22.0, 23.0])
        save_data = isopy.array(data)

        isopy.write_csv(filename('test_io1.csv'), save_data)
        time.sleep(sleep_time)
        read = isopy.read_csv(filename('test_io1.csv'))
        time.sleep(sleep_time)
        self.compare(data, read)

        # With a comment
        isopy.write_csv(filename('test_io2.csv'), save_data, comments='This is a comment')
        time.sleep(sleep_time)
        read = isopy.read_csv(filename('test_io2.csv'))
        time.sleep(sleep_time)
        self.compare(data, read)

        # Several comments and different comment symbol
        isopy.write_csv(filename('test_io3.csv'), save_data, comments=['This is a comment', 'so is this'],
                        comment_symbol='%')
        time.sleep(sleep_time)
        read = isopy.read_csv(filename('test_io3.csv'), comment_symbol='%')
        time.sleep(sleep_time)
        self.compare(data, read)

        # StringIO
        file = io.StringIO()
        isopy.write_csv(file, save_data)
        read = isopy.read_csv(file)
        self.compare(data, read)

        file.seek(10)
        read = isopy.read_csv(file)
        self.compare(data, read)

        # With a comment
        file = io.StringIO()
        isopy.write_csv(file, save_data, comments='This is a comment')
        read = isopy.read_csv(file)
        self.compare(data, read)

        file.seek(10)
        read = isopy.read_csv(file)
        self.compare(data, read)

        # Several comments and different comment symbol
        file = io.StringIO()
        isopy.write_csv(file, save_data, comments=['This is a comment', 'so is this'],
                        comment_symbol='%')
        read = isopy.read_csv(file, comment_symbol='%')
        self.compare(data, read)

        file.seek(10)
        read = isopy.read_csv(file, comment_symbol='%')
        self.compare(data, read)

        # BytesIO & bytes
        file = io.BytesIO()
        isopy.write_csv(file, save_data)
        read = isopy.read_csv(file)
        self.compare(data, read)

        file.seek(10)
        read = isopy.read_csv(file)
        self.compare(data, read)

        file.seek(0)
        filebytes = file.read()
        read = isopy.read_csv(filebytes)
        self.compare(data, read)

        file.seek(0)
        filebytes = file.read()
        filebytes = filebytes.decode('UTF-8')
        filebytes = filebytes.encode('ASCII')
        read = isopy.read_csv(filebytes)
        self.compare(data, read)

        # With a comment
        file = io.BytesIO()
        isopy.write_csv(file, save_data, comments='This is a comment')
        read = isopy.read_csv(file)
        self.compare(data, read)

        file.seek(10)
        read = isopy.read_csv(file)
        self.compare(data, read)

        file.seek(0)
        filebytes = file.read()
        read = isopy.read_csv(filebytes)
        self.compare(data, read)

        file.seek(0)
        filebytes = file.read()
        filebytes = filebytes.decode('UTF-8')
        filebytes = filebytes.encode('ASCII')
        read = isopy.read_csv(filebytes)
        self.compare(data, read)

        # Several comments and different comment symbol
        file = io.BytesIO()
        isopy.write_csv(file, save_data, comments=['This is a comment', 'so is this'],
                        comment_symbol='%')
        read = isopy.read_csv(file, comment_symbol='%')
        self.compare(data, read)

        file.seek(10)
        read = isopy.read_csv(file, comment_symbol='%')
        self.compare(data, read)

        file.seek(0)
        filebytes = file.read()
        read = isopy.read_csv(filebytes, comment_symbol='%')
        self.compare(data, read)

        file.seek(0)
        filebytes = file.read()
        filebytes = filebytes.decode('UTF-8')
        filebytes = filebytes.encode('ASCII')
        read = isopy.read_csv(filebytes, comment_symbol='%')
        self.compare(data, read)

    def compare(self, written, read):
        if isinstance(written, dict):
            assert type(read) is dict
            for key in written.keys():
                assert str(key) in read.keys()
                np.testing.assert_allclose(np.array(read[key], dtype=np.float64), written[key])
        else:
            assert type(read) is list
            np.testing.assert_allclose(np.array(read, dtype=np.float64), written)

    def test_exceptions(self):
        with pytest.raises(TypeError):
            isopy.read_csv(12)

        with pytest.raises(TypeError):
            isopy.write_csv(12, 1)

    def test_keys_in_first_r(self):
        data_d = dict(Pd=[1.0, 2.0, 3.0], Cd=[11.0, np.nan, 13.0], Ru=[21.0, 22.0, 23.0])
        keys = list(data_d.keys())
        values = list(data_d.values())

        # Standard
        file = io.StringIO()
        writer = csv.writer(file, dialect='excel')
        writer.writerow([k for k in keys])
        writer.writerow([v[0] for v in values])
        writer.writerow([v[1] for v in values])
        writer.writerow([v[2] for v in values])

        read = isopy.read_csv(file)
        self.compare(data_d, read)

        read = isopy.read_csv(file, keys_in_first='r')
        self.compare(data_d, read)

        file = io.StringIO()
        isopy.write_csv(file, data_d)

        read = isopy.read_csv(file)
        self.compare(data_d, read)

        read = isopy.read_csv(file, keys_in_first='r')
        self.compare(data_d, read)

        file = io.StringIO()
        isopy.write_csv(file, data_d, keys_in_first='r')

        read = isopy.read_csv(file)
        self.compare(data_d, read)

        read = isopy.read_csv(file, keys_in_first='r')
        self.compare(data_d, read)

        # Comments in the middle
        file = io.StringIO()
        writer = csv.writer(file, dialect='excel')
        writer.writerow([k for k in keys])
        writer.writerow([v[0] for v in values])
        writer.writerow(['# Ignore row',  12, 13])
        writer.writerow([v[1] for v in values])
        writer.writerow([v[2] for v in values])

        read = isopy.read_csv(file)
        self.compare(data_d, read)

        # Uneven row size
        file = io.StringIO()
        writer = csv.writer(file, dialect='excel')
        writer.writerow([k for k in keys])
        writer.writerow([v[0] for v in values])
        writer.writerow([v[1] for v in values[:-1]])
        writer.writerow([v[2] for v in values])

        with pytest.raises(ValueError):
            isopy.read_csv(file)

    def test_keys_in_first_c(self):
        data_d = dict(Pd=[1.0, 2.0, 3.0], Cd=[11.0, np.nan, 13.0], Ru=[21.0, 22.0, 23.0])
        keys = list(data_d.keys())
        values = list(data_d.values())

        # standard
        file = io.StringIO()
        writer = csv.writer(file, dialect='excel')
        writer.writerow([keys[0]] + values[0])
        writer.writerow([keys[1]] + values[1])
        writer.writerow([keys[2]] + values[2])

        read = isopy.read_csv(file)
        self.compare(data_d, read)

        read = isopy.read_csv(file, keys_in_first='c')
        self.compare(data_d, read)

        file = io.StringIO()
        isopy.write_csv(file, data_d, keys_in_first='c')

        read = isopy.read_csv(file)
        self.compare(data_d, read)

        read = isopy.read_csv(file, keys_in_first='c')
        self.compare(data_d, read)

        # Comments in the middle
        file = io.StringIO()
        writer = csv.writer(file, dialect='excel')
        writer.writerow([keys[0]] + values[0])
        writer.writerow(['# Ignore row', 1, 2, 3])
        writer.writerow([keys[1]] + values[1])
        writer.writerow([keys[2]] + values[2])

        read = isopy.read_csv(file)
        self.compare(data_d, read)

        # Uneven row size
        file = io.StringIO()
        writer = csv.writer(file, dialect='excel')
        writer.writerow([keys[0]] + values[0])
        writer.writerow([keys[1]] + values[1][:-1])
        writer.writerow([keys[2]] + values[2])

        with pytest.raises(ValueError):
            isopy.read_csv(file)

    def test_no_keys(self):
        data_l = [[1.0, 11.0, 21.0], [2.0, np.nan, 22.0], [3.0, 13.0, 23.0]]

        # Standard
        file = io.StringIO()
        writer = csv.writer(file, dialect='excel')
        writer.writerow(data_l[0])
        writer.writerow(data_l[1])
        writer.writerow(data_l[2])

        read = isopy.read_csv(file)
        self.compare(data_l, read)

        file = io.StringIO()
        isopy.write_csv(file, data_l)

        read = isopy.read_csv(file)
        self.compare(data_l, read)

        # Comments in the middle
        file = io.StringIO()
        writer = csv.writer(file, dialect='excel')
        writer.writerow(data_l[0])
        writer.writerow(['# Ignore row', 12, 13])
        writer.writerow(data_l[1])
        writer.writerow(data_l[2])

        read = isopy.read_csv(file)
        self.compare(data_l, read)

        # Uneven row size
        file = io.StringIO()
        writer = csv.writer(file, dialect='excel')
        writer.writerow(data_l[0])
        writer.writerow(data_l[1][:-1])
        writer.writerow(data_l[2])

        with pytest.raises(ValueError):
            isopy.read_csv(file)

    def test_no_data(self):
        file = io.StringIO()
        isopy.write_csv(file, [])

        read = isopy.read_csv(file)
        assert read == [[]]

        file = io.StringIO()
        isopy.write_csv(file, [], comments='Empty file')

        read = isopy.read_csv(file)
        assert read == [[]]

    def test_dialect(self):
        # Different dialect
        data_d = dict(Pd=[1.0, 2.0, 3.0], Cd=[11.0, np.nan, 13.0], Ru=[21.0, 22.0, 23.0])
        keys = list(data_d.keys())
        values = list(data_d.values())

        file = io.StringIO()
        writer = csv.writer(file, dialect='excel-tab')
        writer.writerow([k for k in keys])
        writer.writerow([v[0] for v in values])
        writer.writerow([v[1] for v in values])
        writer.writerow([v[2] for v in values])

        read = isopy.read_csv(file, dialect='excel-tab')
        self.compare(data_d, read)

        read = isopy.read_csv(file, dialect=None)
        self.compare(data_d, read)

        file = io.StringIO()
        writer = csv.writer(file, delimiter=';')
        writer.writerow([k for k in keys])
        writer.writerow([v[0] for v in values])
        writer.writerow([v[1] for v in values])
        writer.writerow([v[2] for v in values])

        read = isopy.read_csv(file, dialect=None)
        self.compare(data_d, read)

    def test_nan_values(self):
        for nanstr in isopy.io.NAN_STRINGS:
            data = [[1.0, 11.0, 21.0], [2.0, nanstr, 22.0], [3.0, 13.0, 23.0]]

            file = io.StringIO()
            writer = csv.writer(file, dialect='excel')
            writer.writerow(data[0])
            writer.writerow(data[1])
            writer.writerow(data[2])

            read = isopy.read_csv(file)
            assert np.isnan(read[1][1])

    def test_clipboard(self):
        # Since this just forward to CSV reader we can dont need extensive tests
        # Also this test wont run on Travis

        data = dict(Pd=[1.0, 2.0, 3.0], Cd=[11.0, np.nan, 13.0], Ru=[21.0, 22.0, 23.0])
        save_data = isopy.array(data)

        isopy.write_clipboard(save_data)
        read = isopy.read_clipboard()
        self.compare(data, read)

    def test_truncate(self):
        # Make sure that file like objects are cleared.
        data1 = [[1,2,3], [4,5,6]]
        data2 = [[10, 11], [12, 13], [14, 15]]

        file = io.StringIO()
        isopy.write_csv(file, data1)
        isopy.write_csv(file, data2)
        read = isopy.read_csv(file)
        self.compare(np.array(read, dtype=np.float64), data2)

        file = io.BytesIO()
        isopy.write_csv(file, data1)
        isopy.write_csv(file, data2)
        read = isopy.read_csv(file)
        self.compare(np.array(read, dtype=np.float64), data2)

        # openpyxl truncates BytesIO objects automatically via the use of ZipFile('w')

class Test_xlsx:
    def test_all(self):
        data = dict(Pd=[1.0, 2.0, 3.0], Cd=[11.0, np.nan, 13.0], Ru=[21.0, 22.0, 23.0])
        save_data = isopy.array(data)

        isopy.write_xlsx(filename('test_io1.xlsx'), save_data)
        read = isopy.read_xlsx(filename('test_io1.xlsx'))
        assert type(read) is dict
        assert len(read) == 1
        assert 'sheet1' in read
        self.compare(data, read['sheet1'])

        read = isopy.read_xlsx(filename('test_io1.xlsx'), 'sheet1')
        self.compare(data, read)

        read = isopy.read_xlsx(filename('test_io1.xlsx'), 0)
        self.compare(data, read)

        isopy.write_xlsx(filename('test_io1.xlsx'), test_sheet=save_data)
        read = isopy.read_xlsx(filename('test_io1.xlsx'))
        assert type(read) is dict
        assert len(read) == 1
        assert 'test_sheet' in read
        self.compare(data, read['test_sheet'])

        read = isopy.read_xlsx(filename('test_io1.xlsx'), 'test_sheet')
        self.compare(data, read)

        read = isopy.read_xlsx(filename('test_io1.xlsx'), 0)
        self.compare(data, read)

        # With a comment
        isopy.write_xlsx(filename('test_io2.xlsx'), save_data, comments='This is a comment')
        read = isopy.read_xlsx(filename('test_io2.xlsx'))
        assert type(read) is dict
        assert len(read) == 1
        assert 'sheet1' in read
        self.compare(data, read['sheet1'])

        isopy.write_xlsx(filename('test_io2.xlsx'), test_sheet=save_data, comments='This is a comment')
        read = isopy.read_xlsx(filename('test_io2.xlsx'))
        assert type(read) is dict
        assert len(read) == 1
        assert 'test_sheet' in read
        self.compare(data, read['test_sheet'])

        # Several comments and different comment symbol
        isopy.write_xlsx(filename('test_io3.xlsx'), save_data, comments=['This is a comment', 'so is this'],
                         comment_symbol='%')
        read = isopy.read_xlsx(filename('test_io3.xlsx'), comment_symbol='%')
        assert type(read) is dict
        assert len(read) == 1
        assert 'sheet1' in read
        self.compare(data, read['sheet1'])


        isopy.write_xlsx(filename('test_io3.xlsx'), test_sheet=save_data, comments=['This is a comment', 'so is this'],
                         comment_symbol='%')
        read = isopy.read_xlsx(filename('test_io3.xlsx'), comment_symbol='%')
        assert type(read) is dict
        assert len(read) == 1
        assert 'test_sheet' in read
        self.compare(data, read['test_sheet'])

        # BytesIO and bytes
        file = io.BytesIO()
        isopy.write_xlsx(file, save_data)
        read = isopy.read_xlsx(file)
        assert type(read) is dict
        assert len(read) == 1
        assert 'sheet1' in read
        self.compare(data, read['sheet1'])

        read = isopy.read_xlsx(file, 'sheet1')
        self.compare(data, read)

        file.seek(0)
        filebytes = file.read()
        read = isopy.read_xlsx(filebytes)
        assert type(read) is dict
        assert len(read) == 1
        assert 'sheet1' in read
        self.compare(data, read['sheet1'])

        read = isopy.read_xlsx(filebytes, 'sheet1')
        self.compare(data, read)

        file = io.BytesIO()
        isopy.write_xlsx(file, test_sheet=save_data)
        read = isopy.read_xlsx(file)
        assert type(read) is dict
        assert len(read) == 1
        assert 'test_sheet' in read
        self.compare(data, read['test_sheet'])

        read = isopy.read_xlsx(file, 'test_sheet')
        self.compare(data, read)

        file.seek(0)
        filebytes = file.read()
        read = isopy.read_xlsx(filebytes)
        assert type(read) is dict
        assert len(read) == 1
        assert 'test_sheet' in read
        self.compare(data, read['test_sheet'])

        read = isopy.read_xlsx(filebytes, 'test_sheet')
        self.compare(data, read)

        # With a comment
        file = io.BytesIO()
        isopy.write_xlsx(file, save_data, comments='This is a comment')
        read = isopy.read_xlsx(file)
        assert type(read) is dict
        assert len(read) == 1
        assert 'sheet1' in read
        self.compare(data, read['sheet1'])

        file.seek(0)
        filebytes = file.read()
        read = isopy.read_xlsx(filebytes)
        assert type(read) is dict
        assert len(read) == 1
        assert 'sheet1' in read
        self.compare(data, read['sheet1'])

        file = io.BytesIO()
        isopy.write_xlsx(file, test_sheet=save_data, comments='This is a comment')
        read = isopy.read_xlsx(file)
        assert type(read) is dict
        assert len(read) == 1
        assert 'test_sheet' in read
        self.compare(data, read['test_sheet'])

        file.seek(0)
        filebytes = file.read()
        read = isopy.read_xlsx(filebytes)
        assert type(read) is dict
        assert len(read) == 1
        assert 'test_sheet' in read
        self.compare(data, read['test_sheet'])

        # Several comments and different comment symbol
        file = io.BytesIO()
        isopy.write_xlsx(file, save_data, comments=['This is a comment', 'so is this'],
                         comment_symbol='%')
        read = isopy.read_xlsx(file, comment_symbol='%')
        assert type(read) is dict
        assert len(read) == 1
        assert 'sheet1' in read
        self.compare(data, read['sheet1'])

        file.seek(0)
        filebytes = file.read()
        read = isopy.read_xlsx(filebytes, comment_symbol='%')
        assert type(read) is dict
        assert len(read) == 1
        assert 'sheet1' in read
        self.compare(data, read['sheet1'])

        file = io.BytesIO()
        isopy.write_xlsx(file, test_sheet=save_data, comments=['This is a comment', 'so is this'],
                         comment_symbol='%')
        read = isopy.read_xlsx(file, comment_symbol='%')
        assert type(read) is dict
        assert len(read) == 1
        assert 'test_sheet' in read
        self.compare(data, read['test_sheet'])

        file.seek(0)
        filebytes = file.read()
        read = isopy.read_xlsx(filebytes, comment_symbol='%')
        assert type(read) is dict
        assert len(read) == 1
        assert 'test_sheet' in read
        self.compare(data, read['test_sheet'])

    def test_append(self):
        data1 = dict(Pd=[1.1, 2.2, 3.3], Cd=[11.1, 12.2, 13.3], Ru=[21.1, 22.2, 23.3])
        data2 = dict(Pd=[100.1, 200.2], Cd=[110.1, 120.2], Ru=[210.1, 220.2])

        file = filename('test_io4.xlsx')
        isopy.write_xlsx(file, data1)
        read = isopy.read_xlsx(file)
        assert type(read) is dict
        assert len(read) == 1
        assert 'sheet1' in read
        self.compare(data1, read['sheet1'])

        # appends sheetname
        isopy.write_xlsx(file, test_sheet=data2, append=True)
        read = isopy.read_xlsx(file)
        assert type(read) is dict
        assert len(read) == 2
        assert 'sheet1' in read
        assert 'test_sheet' in read
        self.compare(data1, read['sheet1'])
        self.compare(data2, read['test_sheet'])

        # Should overwrite sheet
        isopy.write_xlsx(file, test_sheet=data1, append=True)
        read = isopy.read_xlsx(file)
        assert type(read) is dict
        assert len(read) == 2
        assert 'sheet1' in read
        assert 'test_sheet' in read
        self.compare(data1, read['sheet1'])
        self.compare(data1, read['test_sheet'])

        # Should overwrite file
        isopy.write_xlsx(file, test_sheet=data2, append=False)
        read = isopy.read_xlsx(file)
        assert type(read) is dict
        assert len(read) == 1
        assert 'test_sheet' in read
        self.compare(data2, read['test_sheet'])

        # BytesIO
        file = io.BytesIO()
        isopy.write_xlsx(file, data1, append=True)
        read = isopy.read_xlsx(file)
        assert type(read) is dict
        assert len(read) == 1
        assert 'sheet1' in read
        self.compare(data1, read['sheet1'])

        # appends sheetname
        isopy.write_xlsx(file, test_sheet=data2, append=True)
        read = isopy.read_xlsx(file)
        assert type(read) is dict
        assert len(read) == 2
        assert 'sheet1' in read
        assert 'test_sheet' in read
        self.compare(data1, read['sheet1'])
        self.compare(data2, read['test_sheet'])

        # Should overwrite sheet
        isopy.write_xlsx(file, test_sheet=data1, append=True)
        read = isopy.read_xlsx(file)
        assert type(read) is dict
        assert len(read) == 2
        assert 'sheet1' in read
        assert 'test_sheet' in read
        self.compare(data1, read['sheet1'])
        self.compare(data1, read['test_sheet'])

        # Should overwrite file
        isopy.write_xlsx(file, test_sheet=data2, append=False)
        read = isopy.read_xlsx(file)
        assert type(read) is dict
        assert len(read) == 1
        assert 'test_sheet' in read
        self.compare(data2, read['test_sheet'])

        # Workbook
        file = openpyxl.Workbook()
        file.remove(file.active) # Remove the automatically created workbook
        isopy.write_xlsx(file, data1)
        read = isopy.read_xlsx(file)
        assert type(read) is dict
        assert len(read) == 1
        assert 'sheet1' in read
        self.compare(data1, read['sheet1'])

        # appends sheetname
        # This should always open in append mode
        isopy.write_xlsx(file, test_sheet=data2)
        read = isopy.read_xlsx(file)
        assert type(read) is dict
        assert len(read) == 2
        assert 'sheet1' in read
        assert 'test_sheet' in read
        self.compare(data1, read['sheet1'])
        self.compare(data2, read['test_sheet'])

        # Should overwrite sheet
        isopy.write_xlsx(file, test_sheet=data1)
        read = isopy.read_xlsx(file)
        assert type(read) is dict
        assert len(read) == 2
        assert 'sheet1' in read
        assert 'test_sheet' in read
        self.compare(data1, read['sheet1'])
        self.compare(data1, read['test_sheet'])

        # Should NOT overwrite file
        isopy.write_xlsx(file, test_sheet=data2, append=False)
        read = isopy.read_xlsx(file)
        assert type(read) is dict
        assert len(read) == 2
        assert 'sheet1' in read
        assert 'test_sheet' in read
        self.compare(data1, read['sheet1'])
        self.compare(data2, read['test_sheet'])

    def compare(self, written, read):
        if isinstance(written, dict):
            assert type(read) is dict
            for key in written.keys():
                assert str(key) in read.keys()
                np.testing.assert_allclose(np.array(read[key], dtype=np.float64), written[key])
        else:
            assert type(read) is list
            np.testing.assert_allclose(np.array(read, dtype=np.float64), written)

    def test_exceptions(self):
        data_d = dict(Pd=[1.1, 2.2, 3.3], Cd=[11.1, np.nan, 13.3], Ru=[21.1, 22.2, 23.3])
        data_ds = core.RefValDict(data_d)
        data_a = isopy.array(data_d)
        data_l = data_a.to_list()

        file = io.BytesIO()
        isopy.write_xlsx(file, data_d)

        with pytest.raises(ValueError):
            isopy.read_xlsx(file, 'nonexistentsheet')

        with pytest.raises(ValueError):
            isopy.read_xlsx(file, 1)

        with pytest.raises(ValueError):
            isopy.read_xlsx(file, keys_in_first = 'b')

    def test_start_at(self):
        data_d1 = dict(Pd=[1.1, 2.2, 3.3], Cd=[11.1, np.nan, 13.3], Ru=[21.1, 22.2, 23.3])
        data_a1 = isopy.array(data_d1)
        data_l1 = data_a1.to_list()

        data_d2 = dict(Rh=[101.1, 102.2], Ag=[1011.1, np.nan], Te=[1021.1, 1022.2])
        data_a2 = isopy.array(data_d2)
        data_l2 = data_a2.to_list()

        data_d3 = dict(Rh=[101.1, 102.2, 3.3], Ag=[1011.1, np.nan, 13.3], Te=[1021.1, 1022.2, 23.3])
        data_a3 = isopy.array(data_d3)
        data_l3 = data_a3.to_list()

        file = io.BytesIO()
        isopy.write_xlsx(file, data_d1)

        read = isopy.read_xlsx(file)
        self.compare(data_d1, read['sheet1'])

        # Skips the header
        read = isopy.read_xlsx(file, start_at='A2')
        self.compare(data_l1, read['sheet1'])

        # Test clear
        isopy.write_xlsx(file, data_d3, append=True, clear=False)
        read = isopy.read_xlsx(file)
        self.compare(data_d3, read['sheet1'])

        isopy.write_xlsx(file, data_d2, append=True)
        read = isopy.read_xlsx(file)
        self.compare(data_d2, read['sheet1'])

        # Write to random cell
        isopy.write_xlsx(file, data_d1)
        isopy.write_xlsx(file, data_d2, start_at='E1', append=True, clear=False)

        read = isopy.read_xlsx(file)
        self.compare(data_d1, read['sheet1'])

        read = isopy.read_xlsx(file, start_at='E1')
        self.compare(data_d2, read['sheet1'])

        # Write to random cell
        isopy.write_xlsx(file, data_d1)
        isopy.write_xlsx(file, data_d2, start_at=(1, 5), append=True, clear=False)

        read = isopy.read_xlsx(file)
        self.compare(data_d1, read['sheet1'])

        read = isopy.read_xlsx(file, start_at='E1')
        self.compare(data_d2, read['sheet1'])

        # Write to random cell
        isopy.write_xlsx(file, data_d1)
        isopy.write_xlsx(file, data_d2, start_at='E1', append=True, clear=False)

        read = isopy.read_xlsx(file)
        self.compare(data_d1, read['sheet1'])

        read = isopy.read_xlsx(file, start_at=(1, 5))
        self.compare(data_d2, read['sheet1'])

    def create_sheet(self, sheetname, data, comments = []):
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)  # Remove the automatically created workbook
        sheet = workbook.create_sheet(sheetname)
        start_r = 1

        for comment in comments:
            sheet.cell(start_r, 1).value = comment
            start_r += 1

        for ri in range(0, len(data)):
            for ci in range(0, len(data[ri])):
                sheet.cell(start_r + ri, ci+1).value = data[ri][ci]

        file = io.BytesIO()
        workbook.save(file)
        return file

    def test_nan_values(self):
        for nanstr in isopy.io.NAN_STRINGS:
            data = [[1.0, 11.0, 21.0], [2.0, nanstr, 22.0], [3.0, 13.0, 23.0]]

            file = self.create_sheet('sheet1', data)
            read = isopy.read_xlsx(file, 'sheet1')
            try:
                assert np.isnan(read[1][1])
            except:
                raise ValueError((nanstr, read))

    def test_keys_in_first_r(self):
        data_d = dict(Pd=[1.0, 2.0, 3.0], Cd=[11.0, np.nan, 13.0], Ru=[21.0, 22.0, 23.0])
        data_l = [list(data_d.keys())] + [[1.0, 2.0, 3.0], [11.0, np.nan, 13.0], [21.0, 22.0, 23.0]]
        data_l = [list(data_d.keys())] + [[1.0, 11.0, 21.0], [2.0, np.nan, 22.0], [3.0, 13.0, 23.0]]

        # Standard
        file = self.create_sheet('sheet1', data_l)

        read = isopy.read_xlsx(file, 'sheet1')
        self.compare(data_d, read)

        read = isopy.read_xlsx(file, 'sheet1', keys_in_first='r')
        self.compare(data_d, read)

        # Comments in the middle
        data_l = data_l[:-1] + [['#Ignore', 33, 34]] + [data_l[-1]]
        file = self.create_sheet('sheet1', data_l)

        read = isopy.read_xlsx(file, 'sheet1')
        self.compare(data_d, read)

    def test_keys_in_first_c(self):
        data_d = dict(Pd=[1.0, 2.0, 3.0], Cd=[11.0, np.nan, 13.0], Ru=[21.0, 22.0, 23.0])
        data_l = [['Pd', 1.0, 2.0, 3.0], ['Cd', 11.0, np.nan, 13.0], ['Ru', 21.0, 22.0, 23.0]]

        # Standard
        file = self.create_sheet('sheet1', data_l)

        read = isopy.read_xlsx(file, 'sheet1')
        self.compare(data_d, read)

        read = isopy.read_xlsx(file, 'sheet1', keys_in_first='c')
        self.compare(data_d, read)

        # Comments in the middle
        data_l = data_l[:-1] + [['#Ignore', 33, 34, 35]] + [data_l[-1]]
        file = self.create_sheet('sheet1', data_l)

        read = isopy.read_xlsx(file, 'sheet1')
        self.compare(data_d, read)

    def test_no_keys(self):
        data_l = [[1.0, 11.0, 21.0], [2.0, np.nan, 22.0], [3.0, 13.0, 23.0]]

        # Standard
        file = self.create_sheet('sheet1', data_l)

        read = isopy.read_xlsx(file, 'sheet1')
        self.compare(data_l, read)

        # Comments in the middle
        data_l2 = data_l[:-1] + [['#Ignore', 33, 34]] + [data_l[-1]]
        file = self.create_sheet('sheet1', data_l2)

        read = isopy.read_xlsx(file, 'sheet1')
        self.compare(data_l, read)

    def test_no_data(self):
        data = []

        file = self.create_sheet('sheet1', data)
        read = isopy.read_xlsx(file, 'sheet1')
        assert read == [[]]

        file = self.create_sheet('sheet1', data, ['#Empty File'])
        read = isopy.read_xlsx(file, 'sheet1')
        assert read == [[]]

# TODO test for pandas
class Test_ToFrom:
    def compare_k(self, read, true):
        assert type(read) == type(true)
        assert read.keys() == true.keys()
        assert read.ndim == true.ndim
        for key in read.keys():
            np.testing.assert_allclose(read[key], true[key])

    def compare_nd(self, read, true):
        assert type(read) == type(true)
        assert read.shape == true.shape
        np.testing.assert_allclose(read, true)

    # CSV + clipboard
    ## array
    def test_a_to_csv(self):
        data = isopy.array(Pd=[1.0, 2.0, 3.0], Cd=[11.0, np.nan, 13.0], Ru=[21.0, 22.0, 23.0])

        for file in [filename('tofrom.csv'), io.StringIO(), io.BytesIO()]:
            data.to_csv(file)
            time.sleep(sleep_time)
            read = isopy.read_csv(file)
            self.compare_k(isopy.array(read), data)

            data.to_csv(file, 'This is a comment', comment_symbol='%')
            time.sleep(sleep_time)
            read = isopy.read_csv(file, comment_symbol='%')
            self.compare_k(isopy.array(read), data)

            data.to_csv(file, keys_in_first='c')
            time.sleep(sleep_time)
            read = isopy.read_csv(file, keys_in_first='c')
            self.compare_k(isopy.array(read), data)

            data.to_csv(file, dialect='excel-tab')
            time.sleep(sleep_time)
            read = isopy.read_csv(file, dialect='excel-tab')
            self.compare_k(isopy.array(read), data)

    def test_a_from_csv(self):
        data = isopy.array(Pd=[1.0, 2.0, 3.0], Cd=[11.0, np.nan, 13.0], Ru=[21.0, 22.0, 23.0])

        for file in [filename('tofrom.csv'), io.StringIO(), io.BytesIO()]:
            for from_csv in [core.IsopyArray.from_csv, isopy.array, isopy.asarray, isopy.asanyarray, isopy.array_from_csv]:
                isopy.write_csv(file, data)
                time.sleep(sleep_time)
                read = from_csv(file)
                self.compare_k(read, data)

                isopy.write_csv(file, data,'This is a comment', comment_symbol='%')
                time.sleep(sleep_time)
                read = from_csv(file, comment_symbol='%')
                self.compare_k(read, data)

                isopy.write_csv(file, data, keys_in_first='c')
                time.sleep(sleep_time)
                read = from_csv(file, keys_in_first='c')
                self.compare_k(read, data)

                isopy.write_csv(file, data, dialect='excel-tab')
                time.sleep(sleep_time)
                read = from_csv(file, dialect='excel-tab')
                self.compare_k(read, data)

        # Test kwargs
        data = isopy.array(Pd=1.0, Cd=np.nan, Ru=21.0)

        file = io.BytesIO()
        isopy.write_csv(file, data)
        read = core.IsopyArray.from_csv(file, ndim=0)
        self.compare_k(read, data)

    def test_a_to_clipboard(self):
        data = isopy.array(Pd=[1.0, 2.0, 3.0], Cd=[11.0, np.nan, 13.0], Ru=[21.0, 22.0, 23.0])


        data.to_clipboard()
        read = isopy.read_clipboard()
        self.compare_k(isopy.array(read), data)

        data.to_clipboard('This is a comment', comment_symbol='%')
        read = isopy.read_clipboard(comment_symbol='%')
        self.compare_k(isopy.array(read), data)

        data.to_clipboard(keys_in_first='c')
        read = isopy.read_clipboard(keys_in_first='c')
        self.compare_k(isopy.array(read), data)

        data.to_clipboard(dialect='excel-tab')
        read = isopy.read_clipboard(dialect='excel-tab')
        self.compare_k(isopy.array(read), data)

    def test_a_from_clipboard(self):
        data = isopy.array(Pd=[1.0, 2.0, 3.0], Cd=[11.0, np.nan, 13.0], Ru=[21.0, 22.0, 23.0])

        for from_clipboard in [core.IsopyArray.from_clipboard, isopy.array_from_clipboard]:
            isopy.write_clipboard(data)
            read = from_clipboard()
            self.compare_k(read, data)

            isopy.write_clipboard(data, 'This is a comment', comment_symbol='%')
            read = from_clipboard(comment_symbol='%')
            self.compare_k(read, data)

            isopy.write_clipboard(data, keys_in_first='c')
            read = from_clipboard(keys_in_first='c')
            self.compare_k(read, data)

            isopy.write_clipboard(data, dialect='excel-tab')
            read = from_clipboard(dialect='excel-tab')
            self.compare_k(read, data)

        for from_csv in [isopy.array, isopy.asarray, isopy.asanyarray]:
            isopy.write_clipboard(data)
            read = from_csv('clipboard')
            self.compare_k(read, data)

            isopy.write_clipboard(data, 'This is a comment', comment_symbol='%')
            read = from_csv('clipboard', comment_symbol='%')
            self.compare_k(read, data)

            isopy.write_clipboard(data, keys_in_first='c')
            read = from_csv('clipboard', keys_in_first='c')
            self.compare_k(read, data)

            isopy.write_clipboard(data, dialect='excel-tab')
            read = from_csv('clipboard', dialect='excel-tab')
            self.compare_k(read, data)

        data = isopy.array(Pd=1.0, Cd=np.nan, Ru=21.0)

        isopy.write_clipboard(data)
        read = core.IsopyArray.from_clipboard(ndim=0)
        self.compare_k(read, data)

    ## scalardict
    def test_ds_to_csv(self):
        data = core.RefValDict(Pd=[1.0, 2.0, 3.0], Cd=[11.0, np.nan, 13.0], Ru=[21.0, 22.0, 23.0])

        for file in [filename('tofrom.csv'), io.StringIO(), io.BytesIO()]:
            data.to_csv(file)
            time.sleep(sleep_time)
            read = isopy.read_csv(file)
            self.compare_k(core.RefValDict(read), data)

            data.to_csv(file, 'This is a comment', comment_symbol='%')
            time.sleep(sleep_time)
            read = isopy.read_csv(file, comment_symbol='%')
            self.compare_k(core.RefValDict(read), data)

            data.to_csv(file, keys_in_first='c')
            time.sleep(sleep_time)
            read = isopy.read_csv(file, keys_in_first='c')
            self.compare_k(core.RefValDict(read), data)

            data.to_csv(file, dialect='excel-tab')
            time.sleep(sleep_time)
            read = isopy.read_csv(file, dialect='excel-tab')
            self.compare_k(core.RefValDict(read), data)

    def test_ds_from_csv(self):
        data = core.RefValDict(Pd=[1.0, 2.0, 3.0], Cd=[11.0, np.nan, 13.0], Ru=[21.0, 22.0, 23.0])

        for file in [filename('tofrom.csv'), io.StringIO(), io.BytesIO()]:
            isopy.write_csv(file, data)
            time.sleep(sleep_time)
            read = core.RefValDict.from_csv(file)
            self.compare_k(read, data)

            isopy.write_csv(file, data, 'This is a comment', comment_symbol='%')
            time.sleep(sleep_time)
            read = core.RefValDict.from_csv(file, comment_symbol='%')
            self.compare_k(read, data)

            isopy.write_csv(file, data, keys_in_first='c')
            time.sleep(sleep_time)
            read = core.RefValDict.from_csv(file, keys_in_first='c')
            self.compare_k(read, data)

            isopy.write_csv(file, data, dialect='excel-tab')
            time.sleep(sleep_time)
            read = core.RefValDict.from_csv(file, dialect='excel-tab')
            self.compare_k(read, data)

        file = io.BytesIO()
        isopy.write_csv(file, data)
        read = core.RefValDict.from_csv(file, default_value=1)
        np.testing.assert_allclose(read.default_value, 1)

    def test_ds_to_clipboard(self):
        data = core.RefValDict(Pd=[1.0, 2.0, 3.0], Cd=[11.0, np.nan, 13.0], Ru=[21.0, 22.0, 23.0])

        data.to_clipboard()
        read = isopy.read_clipboard()
        self.compare_k(core.RefValDict(read), data)

        data.to_clipboard('This is a comment', comment_symbol='%')
        read = isopy.read_clipboard(comment_symbol='%')
        self.compare_k(core.RefValDict(read), data)

        data.to_clipboard(keys_in_first='c')
        read = isopy.read_clipboard(keys_in_first='c')
        self.compare_k(core.RefValDict(read), data)

        data.to_clipboard(dialect='excel-tab')
        read = isopy.read_clipboard(dialect='excel-tab')
        self.compare_k(core.RefValDict(read), data)

    def test_ds_from_clipboard(self):
        data = core.RefValDict(Pd=[1.0, 2.0, 3.0], Cd=[11.0, np.nan, 13.0], Ru=[21.0, 22.0, 23.0])

        isopy.write_clipboard(data)
        read = core.RefValDict.from_clipboard()
        self.compare_k(read, data)

        isopy.write_clipboard(data, 'This is a comment', comment_symbol='%')
        read = core.RefValDict.from_clipboard(comment_symbol='%')
        self.compare_k(read, data)

        isopy.write_clipboard(data, keys_in_first='c')
        read = core.RefValDict.from_clipboard(keys_in_first='c')
        self.compare_k(read, data)

        isopy.write_clipboard(data, dialect='excel-tab')
        read = core.RefValDict.from_clipboard(dialect='excel-tab')
        self.compare_k(read, data)

        # Kwargs
        isopy.write_clipboard(data)
        read = core.RefValDict.from_clipboard(default_value=1)
        np.testing.assert_allclose(read.default_value, 1)

    ## ndarray
    def test_nd_from_csv(self):
        data = np.array([[1.0, 11.0, 21.0], [2.0, np.nan, 22.0], [3.0, 13.0, 23.0]])

        for file in [filename('tofrom.csv'), io.StringIO(), io.BytesIO()]:
            isopy.write_csv(file, data)
            time.sleep(sleep_time)
            read = isopy.asanyarray(file)
            self.compare_nd(read, data)

            isopy.write_csv(file, 'This is a comment', comment_symbol='%')
            time.sleep(sleep_time)
            read = isopy.asanyarray(file, comment_symbol='%')
            self.compare_nd(read, data)

            isopy.write_csv(file, dialect='excel-tab')
            time.sleep(sleep_time)
            read = isopy.asanyarray(file, dialect='excel-tab')
            self.compare_nd(read, data)

    def test_nd_from_clipboard(self):
        data = np.array([[1.0, 11.0, 21.0], [2.0, np.nan, 22.0], [3.0, 13.0, 23.0]])

        file='clipboard'
        isopy.write_clipboard(data, data)
        read = isopy.asanyarray(file)
        self.compare_nd(read, data)

        isopy.write_clipboard(data, 'This is a comment', comment_symbol='%')
        read = isopy.asanyarray(file, comment_symbol='%')
        self.compare_nd(read, data)

        isopy.write_clipboard(data, dialect='excel-tab')
        read = isopy.asanyarray(file, dialect='excel-tab')
        self.compare_nd(read, data)

    # XLSX
    ## array
    def test_a_to_xlsx(self):
        data1 = isopy.array(Pd=[1.0, 2.0, 3.0], Cd=[11.0, np.nan, 13.0], Ru=[21.0, 22.0, 23.0])
        data2 = data1*100

        for file in [filename('tofrom.xlsx'), io.BytesIO()]:
            data1.to_xlsx(file)
            read = isopy.read_xlsx(file, 'sheet1')
            self.compare_k(isopy.array(read), data1)

            data1.to_xlsx(file, 'sheet2', 'This is a comment', comment_symbol='%')
            read = isopy.read_xlsx(file, 'sheet2', comment_symbol='%')
            self.compare_k(isopy.array(read), data1)

            data1.to_xlsx(file, keys_in_first='c', start_at='E1')
            read = isopy.read_xlsx(file, 'sheet1', keys_in_first='c', start_at='E1')
            self.compare_k(isopy.array(read), data1)

            data1.to_xlsx(file)
            data2.to_xlsx(file, 'sheet2', append=True)
            read = isopy.read_xlsx(file, 'sheet1')
            self.compare_k(isopy.array(read), data1)
            read = isopy.read_xlsx(file, 'sheet2')
            self.compare_k(isopy.array(read), data2)

            data1.to_xlsx(file)
            data2.to_xlsx(file, 'sheet2', append=True, start_at='E1', clear=False)
            read = isopy.read_xlsx(file, 'sheet1')
            self.compare_k(isopy.array(read), data1)
            read = isopy.read_xlsx(file, 'sheet2', start_at='E1')
            self.compare_k(isopy.array(read), data2)

    def test_a_from_xlsx(self):
        data1 = isopy.array(Pd=[1.0, 2.0, 3.0], Cd=[11.0, np.nan, 13.0], Ru=[21.0, 22.0, 23.0])
        data2 = data1 * 100

        for file in [filename('tofrom.xlsx'), io.BytesIO()]:
            for from_xlsx in [core.IsopyArray.from_xlsx, isopy.array, isopy.asarray, isopy.asanyarray, isopy.array_from_xlsx]:
                isopy.write_xlsx(file, data1)
                read = from_xlsx(file, sheetname='sheet1')
                self.compare_k(read, data1)

                isopy.write_xlsx(file, data1, comments='This is a comment', comment_symbol='%')
                read = from_xlsx(file, comment_symbol='%', sheetname='sheet1')
                self.compare_k(read, data1)

                isopy.write_xlsx(file, data1, keys_in_first='c')
                read = from_xlsx(file, keys_in_first='c', sheetname='sheet1')
                self.compare_k(read, data1)

                isopy.write_xlsx(file, data1)
                isopy.write_xlsx(file, append=True, sheet2 = data2)
                read = from_xlsx(file, sheetname='sheet1')
                self.compare_k(read, data1)
                read = from_xlsx(file, sheetname='sheet2')
                self.compare_k(read, data2)

                isopy.write_xlsx(file, data1)
                isopy.write_xlsx(file, data2, append=True, start_at='E1', clear=False)
                read = from_xlsx(file, sheetname='sheet1')
                self.compare_k(read, data1)
                read = from_xlsx(file, start_at='E1', sheetname='sheet1')
                self.compare_k(read, data2)

        # Test kwargs
        data1 = isopy.array(Pd=1.0, Cd=np.nan, Ru=21.0)

        file = io.BytesIO()
        isopy.write_xlsx(file, data1)
        read = core.IsopyArray.from_xlsx(file, ndim=0, sheetname='sheet1')
        self.compare_k(read, data1)

    ## scalardict
    def test_ds_to_xlsx(self):
        data1 = core.RefValDict(Pd=[1.0, 2.0, 3.0], Cd=[11.0, np.nan, 13.0], Ru=[21.0, 22.0, 23.0])
        data2 = np.multiply(data1, 100)

        for file in [filename('tofrom.xlsx'), io.BytesIO()]:
            data1.to_xlsx(file)
            read = isopy.read_xlsx(file, 'sheet1')
            self.compare_k(core.RefValDict(read), data1)

            data1.to_xlsx(file, 'sheet2', 'This is a comment', comment_symbol='%')
            read = isopy.read_xlsx(file, 'sheet2', comment_symbol='%')
            self.compare_k(core.RefValDict(read), data1)

            data1.to_xlsx(file, keys_in_first='c', start_at='E1')
            read = isopy.read_xlsx(file, 'sheet1', keys_in_first='c', start_at='E1')
            self.compare_k(core.RefValDict(read), data1)

            data1.to_xlsx(file)
            data2.to_xlsx(file, 'sheet2', append=True)
            read = isopy.read_xlsx(file, 'sheet1')
            self.compare_k(core.RefValDict(read), data1)
            read = isopy.read_xlsx(file, 'sheet2')
            self.compare_k(core.RefValDict(read), data2)

            data1.to_xlsx(file)
            data2.to_xlsx(file, append=True, start_at='E1', clear=False)
            read = isopy.read_xlsx(file, 'sheet1')
            self.compare_k(core.RefValDict(read), data1)
            read = isopy.read_xlsx(file, 'sheet1', start_at='E1')
            self.compare_k(core.RefValDict(read), data2)

    def test_ds_from_xlsx(self):
        data1 = core.RefValDict(Pd=[1.0, 2.0, 3.0], Cd=[11.0, np.nan, 13.0], Ru=[21.0, 22.0, 23.0])
        data2 = np.multiply(data1, 100)

        for file in [filename('tofrom.xlsx'), io.BytesIO()]:
                isopy.write_xlsx(file, data1)
                read = core.RefValDict.from_xlsx(file, sheetname='sheet1')
                self.compare_k(read, data1)

                isopy.write_xlsx(file, data1, comments='This is a comment', comment_symbol='%')
                read = core.RefValDict.from_xlsx(file, comment_symbol='%', sheetname='sheet1')
                self.compare_k(read, data1)

                isopy.write_xlsx(file, data1, keys_in_first='c')
                read = core.RefValDict.from_xlsx(file, keys_in_first='c', sheetname='sheet1')
                self.compare_k(read, data1)

                isopy.write_xlsx(file, data1)
                isopy.write_xlsx(file, append=True, sheet2=data2)
                read = core.RefValDict.from_xlsx(file, sheetname='sheet1')
                self.compare_k(read, data1)
                read = core.RefValDict.from_xlsx(file, sheetname='sheet2')
                self.compare_k(read, data2)

                isopy.write_xlsx(file, data1)
                isopy.write_xlsx(file, data2, append=True, start_at='E1', clear=False)
                read = core.RefValDict.from_xlsx(file, sheetname='sheet1')
                self.compare_k(read, data1)
                read = core.RefValDict.from_xlsx(file, start_at='E1', sheetname='sheet1')
                self.compare_k(read, data2)

        # Test kwargs
        file = io.BytesIO()
        isopy.write_xlsx(file, data1)
        read = core.RefValDict.from_xlsx(file, default_value=1, sheetname='sheet1')
        np.testing.assert_allclose(read.default_value, 1)

    ## ndarray
    def test_nd_from_csv(self):
        data1 = np.array([[1.0, 11.0, 21.0], [2.0, np.nan, 22.0], [3.0, 13.0, 23.0]])
        data2 = data1 * 100

        for file in [filename('tofrom.xlsx'), io.BytesIO()]:
            isopy.write_xlsx(file, data1)
            read = isopy.asanyarray(file, sheetname='sheet1')
            self.compare_nd(read, data1)

            isopy.write_xlsx(file, data1, comments='This is a comment', comment_symbol='%')
            read = isopy.asanyarray(file, comment_symbol='%', sheetname='sheet1')
            self.compare_nd(read, data1)

            isopy.write_xlsx(file, data1)
            isopy.write_xlsx(file, append=True, sheet2=data2)
            read = isopy.asanyarray(file, sheetname='sheet1')
            self.compare_nd(read, data1)
            read = isopy.asanyarray(file, sheetname='sheet2')
            self.compare_nd(read, data2)

            isopy.write_xlsx(file, data1)
            isopy.write_xlsx(file, data2, append=True, start_at='E1', clear=False,)
            read = isopy.asanyarray(file, sheetname='sheet1')
            self.compare_nd(read, data1)
            read = isopy.asanyarray(file, start_at='E1', sheetname='sheet1')
            self.compare_nd(read, data2)


# TODO check meta data
# TODO check that the measurements are right
# TODO check cycles, time
class Test_exp:
    def test_read(self):
        file = filename('palladium.exp')
        self.palladium_exp(file)

        with open(file, 'rb') as fileio:
            text = fileio.read()

        self.palladium_exp(text)
        self.palladium_exp(io.BytesIO(text))

        text = text.decode('UTF-8')
        self.palladium_exp(io.StringIO(text))


        with pytest.raises(TypeError):
            self.palladium_exp(12)

    def palladium_exp(self, file):
        data = isopy.read_exp(file)

        assert type(data.info) is dict
        assert data.info['Sample ID'] == 'blk'
        assert data.info['Operator'] == 'ME'
        assert data.info['Instrument'] == 'TIGER'

        assert type(data.cycle) is list
        assert data.cycle == [i for i in range(1, 61)]

        assert type(data.time) is list
        assert len(data.time) == len(data.cycle)
        assert isinstance(data.time[0], datetime.datetime)
        assert data.time[0] == datetime.datetime.strptime('14:49:51:318', '%H:%M:%S:%f')
        assert data.time[-1] == datetime.datetime.strptime('14:53:58:927', '%H:%M:%S:%f')

        assert type(data.measurements) is dict
        assert len(data.measurements) == 1
        assert 1 in data.measurements

        assert isinstance(data.measurements[1], core.IsopyArray)
        for key in '101ru 102pd 104pd 105pd 106pd 108pd 110pd 111cd'.split():
            assert key in data.measurements[1].keys

        np.testing.assert_allclose(data.measurements[1]['102pd'][0], 2.9360921148728890e-004)
        np.testing.assert_allclose(data.measurements[1]['110pd'][-1], 3.5484434560084858e-003)

    def test_renamer(self):
        renamer = {'102Pd': '102Ru'}

        data = isopy.read_exp(filename('palladium.exp'), renamer)
        assert '102Pd' not in data.measurements[1].keys
        assert '102Ru' in data.measurements[1].keys

        def renamer(key):
            if key == '110Pd':
                return '110Cd'
            else:
                return key

        data = isopy.read_exp(filename('palladium.exp'), renamer)
        assert '110Pd' not in data.measurements[1].keys
        assert '110Cd' in data.measurements[1].keys